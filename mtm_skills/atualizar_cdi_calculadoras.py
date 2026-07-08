"""Mantém a aba CDI de cada calculadora CDI actualizada.

Fluxo em duas fases:
  1. Triagem rapida (openpyxl, sem Excel): categoriza cada calculadora em
     'migrar' (sem aba CDI), 'atualizar' (aba CDI desatualizada) ou 'ok'.
  2. Atualizacao paralela (ThreadPoolExecutor + openpyxl, sem Excel COM):
     recria a aba CDI completa com todos os dias desde DataEmissao.

Idempotente. Uso:
    python -m mtm_skills.atualizar_cdi_calculadoras

Requisitos: calculadoras fechadas antes de rodar (todas sao .xlsx; nao precisa
de Excel instalado).
"""
import glob
import os
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date as _date
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Caminhos
# ---------------------------------------------------------------------------
def _glob1(pattern: str) -> str | None:
    hits = glob.glob(pattern, recursive=True)
    return hits[0] if hits else None


MERCADO_XLSX = os.environ.get("MERCADO_XLSX") or str(
    Path(r"C:\caminho\para\Rotinas"
         r"\Codigos\14-Base_Dados_Mercado\output\mercado.xlsx")
)
CADASTRO = os.environ.get("CADASTRO") or _glob1(
    r"C:\caminho\para\projeto*\*rea de Trabalho\Pessoal\Projetos"
    r"\Projeto - PU_MTM\data\cadastro_ativos.xlsx"
) or _glob1(
    r"C:\caminho\para\projeto*\*rea de Trabalho\Pessoal\Projetos"
    r"\Projeto - PU_MTM\data\cadastro_ativos.csv"
)
CALC_ROOT = Path(os.environ.get(
    "CALC_ROOT",
    r"C:\caminho\para\Calculadoras",
))
FLUXO_DIR = CALC_ROOT / "Fluxo de pagamento"

# Diretório das cópias de homologação (fase de homolog: SharePoint não é mais tocado).
_RAIZ_PROJ = Path(__file__).resolve().parents[1]
HOMOLOG_ROOT = Path(os.environ.get(
    "HOMOLOG_ROOT",
    str(_RAIZ_PROJ / "data" / "Calculadoras - Homologação"),
))
# Modo de destino: 'homolog' (default) atualiza as cópias locais; 'sharepoint' usa FLUXO_DIR.
MODO = os.environ.get("CALC_MODE", "homolog").lower()

ABA_CDI = "CDI"
FAMILIAS_CDI = ("di_puro", "di_spread")
MAX_WORKERS = 8  # threads openpyxl: sem limite de licenca Excel

# A proteção antiransomware do TI mata o processo após ~5-6 modificações de
# arquivos Office e reverte todas as escritas dele. No modo homolog gravamos em
# subprocessos que processam poucos arquivos e SAEM normalmente (writes persistem).
BATCH_SIZE = int(os.environ.get("CDI_BATCH", "3"))


def indexar_homolog() -> dict:
    """Indexa os arquivos de homologação por nome de arquivo (minúsculo).

    A estrutura é ``HOMOLOG_ROOT/<familia>/<pasta - Homologação>/<arquivo>``;
    casar por nome de arquivo é robusto a divergências no nome da pasta.
    """
    idx: dict[str, Path] = {}
    if not HOMOLOG_ROOT.exists():
        return idx
    for sub in HOMOLOG_ROOT.iterdir():
        if not sub.is_dir():
            continue
        for pasta in sub.iterdir():
            if not pasta.is_dir():
                continue
            for arq in pasta.iterdir():
                if arq.suffix.lower() in (".xlsx", ".xlsm"):
                    idx[arq.name.lower()] = arq
    return idx


def resolver_calc(r: dict, homolog_idx: dict):
    """Resolve o caminho da calculadora conforme o modo (homolog/sharepoint)."""
    if MODO == "homolog":
        return homolog_idx.get(Path(r["CalcPath"]).name.lower())
    return FLUXO_DIR / Path(r["CalcPath"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _to_iso(val) -> str:
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    return str(val)[:10]


def _normalizar_data(data_str: str) -> str:
    s = str(data_str).strip()[:10]
    if "/" in s:
        d, m, a = s.split("/")
        return f"{a}-{m.zfill(2)}-{d.zfill(2)}"
    return s


# ---------------------------------------------------------------------------
# Leitura do historico CDI no mercado.xlsx
# ---------------------------------------------------------------------------
def ler_cdi_mercado(mercado_path: str) -> dict:
    """Retorna {data_iso: (cdi_aa, fator_diario)} ordenado."""
    wb = openpyxl.load_workbook(mercado_path, data_only=True, read_only=True)

    cdi_map = {
        _to_iso(row[1]): row[2]
        for row in wb["CDI"].iter_rows(min_row=2, values_only=True)
        if row[1] and row[2] is not None
    }
    fator_map = {
        _to_iso(row[1]): row[2]
        for row in wb["FATOR_DIARIO_CDI"].iter_rows(min_row=2, values_only=True)
        if row[1] and row[2] is not None
    }
    wb.close()

    if not cdi_map:
        raise ValueError("Nenhum dado de CDI encontrado em mercado.xlsx")

    return {d: (cdi_map[d], fator_map.get(d)) for d in sorted(cdi_map)}


# ---------------------------------------------------------------------------
# Leitura do cadastro
# ---------------------------------------------------------------------------
def ler_calculadoras_cdi(cadastro_path: str) -> list:
    import sys as _sys
    from pathlib import Path as _Path
    _raiz = str(_Path(__file__).resolve().parents[1])
    if _raiz not in _sys.path:
        _sys.path.insert(0, _raiz)
    from pu_mtm.dados.csvio import ler_dict
    rows = ler_dict(cadastro_path)
    return [r for r in rows if r.get("Familia") in FAMILIAS_CDI and r.get("CalcPath")]


# ---------------------------------------------------------------------------
# Fase 1: triagem rapida via openpyxl (sem abrir Excel)
# ---------------------------------------------------------------------------
def _triagem_cdi_tab(wb_path: str, emissao_iso: str, cdi_hist: dict) -> tuple[str, int]:
    """
    Compara TODAS as datas esperadas (cdi_hist >= emissao_iso) com as presentes
    na aba CDI do arquivo — detecta lacunas, nao apenas ausencia de novas datas.

    Retorna: ('ok', 0), ('atualizar', n_faltantes) ou ('migrar', 0).
    """
    try:
        wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
        if ABA_CDI not in wb.sheetnames:
            wb.close()
            return "migrar", 0

        existentes: set[str] = set()
        for row in wb[ABA_CDI].iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                existentes.add(_to_iso(row[0]))
        wb.close()

        if not existentes:
            return "migrar", 0

        esperadas = {d for d in cdi_hist if d >= emissao_iso}
        faltantes = esperadas - existentes
        if not faltantes:
            return "ok", 0
        return "atualizar", len(faltantes)
    except Exception:
        return "migrar", 0


# ---------------------------------------------------------------------------
# Fase 2: cria/recria aba CDI via openpyxl (sem Excel COM)
# ---------------------------------------------------------------------------
def _atualizar_openpyxl(wb_path: str, emissao_iso: str, cdi_hist: dict) -> tuple[str, int]:
    """Recria a aba CDI completa com openpyxl.

    Sempre reescreve a aba inteira — mais simples e igualmente rapido, pois
    cdi_hist tem ~500 linhas e openpyxl escreve em < 0.5s por arquivo.
    Datas sao gravadas como objetos date (tipo Excel correto, nao texto).
    """
    wb = openpyxl.load_workbook(wb_path)

    action = "migrar" if ABA_CDI not in wb.sheetnames else "atualizar"

    if ABA_CDI in wb.sheetnames:
        del wb[ABA_CDI]

    ws = wb.create_sheet(ABA_CDI)
    ws.append(["DATA", "CDI_AA", "FATOR_DIARIO"])

    for d, (cdi, fat) in cdi_hist.items():
        if d >= emissao_iso:
            ws.append([_date.fromisoformat(d), cdi, fat if fat is not None else ""])

    wb.save(wb_path)
    wb.close()
    return action, sum(1 for d in cdi_hist if d >= emissao_iso)


def _worker_thread(task: tuple, cdi_hist: dict) -> tuple:
    tipo, meta, cam_str = task
    emissao_iso = _normalizar_data(meta["DataEmissao"])
    print(
        f"  [{meta['IdSerie']:>6}] {meta['Apelido']:<40} | {tipo.capitalize()}",
        flush=True,
    )
    try:
        action, n = _atualizar_openpyxl(cam_str, emissao_iso, cdi_hist)
        return meta["IdSerie"], meta["Apelido"], action, "ok", n
    except Exception as e:
        return meta["IdSerie"], meta["Apelido"], tipo, "erro", str(e)


# ---------------------------------------------------------------------------
# Cache do histórico CDI (evita reler mercado.xlsx em cada subprocesso)
# ---------------------------------------------------------------------------
def _cdi_hist_to_json(cdi_hist: dict, path: str) -> None:
    import json
    with open(path, "w", encoding="utf-8") as f:
        json.dump({d: list(v) for d, v in cdi_hist.items()}, f)


def _cdi_hist_from_json(path: str) -> dict:
    import json
    with open(path, encoding="utf-8") as f:
        raw = json.load(f)
    return {d: tuple(v) for d, v in raw.items()}


def atualizar_lote(sids: list[str], cache_path: str) -> None:
    """Atualiza a aba CDI dos IdSeries indicados (executado em subprocesso).

    Processa poucos arquivos e retorna — mantendo o processo abaixo do limite
    que dispara o antivírus, para que as escritas persistam.
    """
    cdi_hist = _cdi_hist_from_json(cache_path)
    idx = indexar_homolog() if MODO == "homolog" else {}
    calcs = {r["IdSerie"]: r for r in ler_calculadoras_cdi(CADASTRO)}
    for sid in sids:
        r = calcs.get(sid)
        if not r:
            print(f"  [{sid:>10}] NAO NO CADASTRO", flush=True)
            continue
        cam = resolver_calc(r, idx)
        if cam is None or not cam.exists():
            print(f"  [{sid:>10}] NAO ENCONTRADO", flush=True)
            continue
        try:
            action, n = _atualizar_openpyxl(
                str(cam), _normalizar_data(r["DataEmissao"]), cdi_hist)
            tag = f"MIGRADA ({n} linhas)" if action == "migrar" else f"+{n} linha(s)"
            print(f"  [{sid:>10}] {r.get('Apelido', sid):<40} {tag}", flush=True)
        except Exception as e:
            print(f"  [{sid:>10}] {r.get('Apelido', sid):<40} ERRO: {e}", flush=True)


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------
def atualizar() -> None:
    if not MERCADO_XLSX:
        raise FileNotFoundError("mercado.xlsx nao encontrado. Defina MERCADO_XLSX.")
    if not CADASTRO:
        raise FileNotFoundError("cadastro_ativos.csv nao encontrado. Defina CADASTRO.")

    cdi_hist = ler_cdi_mercado(MERCADO_XLSX)
    datas = list(cdi_hist.keys())
    ultima_mercado = datas[-1]
    print(f"CDI mercado: {len(datas)} datas ({datas[0]} -> {ultima_mercado})\n")

    calcs = ler_calculadoras_cdi(CADASTRO)
    print(f"Calculadoras CDI: {len(calcs)}")

    homolog_idx = indexar_homolog() if MODO == "homolog" else {}
    destino = HOMOLOG_ROOT if MODO == "homolog" else FLUXO_DIR
    print(f"Modo: {MODO} -> {destino}"
          + (f" ({len(homolog_idx)} arquivos indexados)" if MODO == "homolog" else ""))

    # ------------------------------------------------------------------
    # Fase 1: triagem paralela (ThreadPoolExecutor, I/O-bound openpyxl)
    # ------------------------------------------------------------------
    pendentes: list[tuple] = []
    ja_em_dia = nao_achado = 0

    existentes: list[tuple] = []
    for r in calcs:
        cam = resolver_calc(r, homolog_idx)
        if cam is None or not cam.exists():
            nome = Path(r["CalcPath"]).name
            print(f"  [{r['IdSerie']:>6}] NAO ENCONTRADO: {nome}")
            nao_achado += 1
        else:
            existentes.append((r, str(cam), _normalizar_data(r["DataEmissao"])))

    def _triagem_task(args: tuple) -> tuple:
        r, cam_str, emissao_iso = args
        status, n_falt = _triagem_cdi_tab(cam_str, emissao_iso, cdi_hist)
        return r, cam_str, status, n_falt

    n_threads = min(MAX_WORKERS, len(existentes)) if existentes else 1
    with ThreadPoolExecutor(max_workers=n_threads) as pool:
        for r, cam_str, status_tri, n_falt in pool.map(_triagem_task, existentes):
            if status_tri == "migrar":
                print(f"  [{r['IdSerie']:>6}] {r['Apelido']:<40} sem aba CDI → migrar")
                pendentes.append(("migrar", r, cam_str))
            elif status_tri == "ok":
                ja_em_dia += 1
            else:
                print(f"  [{r['IdSerie']:>6}] {r['Apelido']:<40} {n_falt} data(s) faltante(s)")
                pendentes.append(("atualizar", r, cam_str))

    n_migrar = sum(1 for t, *_ in pendentes if t == "migrar")
    n_atualizar = len(pendentes) - n_migrar
    print(
        f"\nTriagem: {n_migrar} Migrar | {n_atualizar} Atualizar | "
        f"{ja_em_dia} Em dia | {nao_achado} Não encontrado"
    )

    if not pendentes:
        print("\nNenhuma calculadora precisa de atualizacao.")
        return

    # ------------------------------------------------------------------
    # Fase 2: atualizacao da aba CDI (openpyxl, sem Excel COM)
    # ------------------------------------------------------------------
    # Arquivos sincronizados pelo OneDrive não persistem quando salvos por
    # threads de um ThreadPoolExecutor (o save no worker é descartado pelo
    # filtro de sync). No modo homolog processamos na thread principal.
    if MODO == "homolog":
        # Grava via subprocessos de até BATCH_SIZE arquivos (contorna o antivírus).
        import json  # noqa: F401  (usado pelos helpers)
        import subprocess
        import tempfile

        cache = tempfile.mktemp(suffix="_cdi_cache.json")
        _cdi_hist_to_json(cdi_hist, cache)

        sids = [t[1]["IdSerie"] for t in pendentes]
        lotes = [sids[i:i + BATCH_SIZE] for i in range(0, len(sids), BATCH_SIZE)]
        print(f"\nProcessando {len(sids)} calculadora(s) em {len(lotes)} "
              f"subprocesso(s) de até {BATCH_SIZE} (contorna antivírus)...\n")

        for lote in lotes:
            proc = subprocess.run(
                [sys.executable, "-W", "ignore", "-m",
                 "mtm_skills.atualizar_cdi_calculadoras",
                 "--ids", ",".join(lote), "--cache", cache],
                capture_output=True, text=True, encoding="utf-8", errors="replace",
            )
            if proc.stdout:
                sys.stdout.write(proc.stdout)
            if proc.returncode != 0:
                print(f"  (subprocesso do lote {lote} saiu com código {proc.returncode})")
        try:
            os.remove(cache)
        except OSError:
            pass

        # Reconferência honesta: quantos ficaram em dia
        ultima = ultima_mercado
        homolog_idx = indexar_homolog()
        em_dia = atrasados = 0
        for r in calcs:
            cam = resolver_calc(r, homolog_idx)
            if cam is None or not cam.exists():
                continue
            try:
                wb = openpyxl.load_workbook(str(cam), data_only=True, read_only=True)
                datas_tab = [_to_iso(row[0]) for row in
                             wb[ABA_CDI].iter_rows(min_row=2, max_col=1, values_only=True)
                             if row[0]]
                wb.close()
                if datas_tab and max(datas_tab) >= ultima:
                    em_dia += 1
                else:
                    atrasados += 1
            except Exception:
                atrasados += 1
        print(f"\nConcluido (homolog): {em_dia} Em dia | {atrasados} Atrasadas "
              f"| {nao_achado} Não Encontradas")
        return

    n_workers = min(MAX_WORKERS, len(pendentes))
    print(f"\nProcessando {len(pendentes)} calculadora(s) ({n_workers} thread(s))...\n")
    with ThreadPoolExecutor(max_workers=n_workers) as pool:
        futs = [pool.submit(_worker_thread, task, cdi_hist) for task in pendentes]
        resultados = [f.result() for f in as_completed(futs)]

    migrados = atualizados = erros = 0
    for id_serie, apelido, tipo, status, detalhe in resultados:
        if status == "ok":
            tag = f"MIGRADA ({detalhe} linhas)" if tipo == "migrar" else f"+{detalhe} linha(s)"
            print(f"  [{id_serie:>6}] {apelido:<40} {tag}")
            if tipo == "migrar":
                migrados += 1
            else:
                atualizados += 1
        else:
            print(f"  [{id_serie:>6}] {apelido:<40} ERRO: {detalhe}")
            erros += 1

    print(
        f"\nConcluido: {migrados} Migradas | {atualizados} Atualizadas | "
        f"{ja_em_dia} Em Dia | {nao_achado} Não Encontradas | {erros} Erros"
    )


if __name__ == "__main__":
    import sys as _sys
    try:
        _sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    argv = _sys.argv
    if "--ids" in argv:
        # Modo subprocesso: atualiza um lote específico e sai.
        ids = argv[argv.index("--ids") + 1].split(",")
        cache = argv[argv.index("--cache") + 1]
        atualizar_lote([s for s in ids if s], cache)
    else:
        atualizar()

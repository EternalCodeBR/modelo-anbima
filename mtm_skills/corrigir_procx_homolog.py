"""Corrige a coluna de taxa (PROCX/XLOOKUP) das calculadoras de homologação.

Estratégia fiel: para cada célula da coluna de taxa, lê a fórmula do ORIGINAL
no SharePoint e substitui apenas a chamada ``SUMIFS([n]SERIE_VALOR!...)`` por
``XLOOKUP(<mesma célula de data>, CDI!A:A, CDI!B:B)``, preservando todo o resto
(offset da data, divisão ``/100`` e envelope ``IF(...)``). Células sem SUMIFS
(valores fixos/seed) são copiadas verbatim.

Assim o homolog reproduz exatamente o cálculo do SharePoint, mas lendo a aba CDI
local em vez do BaseDadosMercado externo.

Os arquivos do SharePoint são apenas LIDOS (via cópia temporária) — nunca
alterados. Só os arquivos em ``data/Calculadoras - Homologação`` são reescritos.

Uso:
    python -m mtm_skills.corrigir_procx_homolog --dry-run   # mostra sem gravar
    python -m mtm_skills.corrigir_procx_homolog             # aplica
    python -m mtm_skills.corrigir_procx_homolog --todos     # processa os 23
"""
import re
import shutil
import sys
import tempfile
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

_RAIZ = Path(__file__).resolve().parents[1]
if str(_RAIZ) not in sys.path:
    sys.path.insert(0, str(_RAIZ))

import openpyxl
from openpyxl.utils import get_column_letter

from pu_mtm.app import config
from pu_mtm.dados.csvio import ler_dict

HOMO_ROOT = _RAIZ / "data" / "Calculadoras - Homologação"
FAMILIAS_CDI = {"di_puro", "di_spread"}
NOMES_TAXA = ["Taxa_DI", "DI Over", "DI"]
NOMES_DATA = ["Data", "DATA", "data"]
ABAS_IGNORAR = {"CDI", "Feriados", "obs"}

# IdSeries com PROCX incorreto (resultado da verificação).
IDS_CORRIGIR = [
    "624", "629", "638", "639", "734", "667", "723",  # offset dia-anterior perdido
    "612",           # referência quebrada (apontava p/ CDI!A)
    "123321123",     # /100 perdido
    "763",           # IF + /100 perdidos (misto)
    "691",           # IF + /100 + coluna de data (Ativo DI+Spread B ERRO)
    "765",           # offset dia-anterior perdido
]


# ---------------------------------------------------------------------------
# Transformação SUMIFS -> XLOOKUP
# ---------------------------------------------------------------------------
def _match_paren(s: str, open_idx: int) -> int:
    """Índice do ')' que fecha o '(' em open_idx."""
    depth = 0
    for i in range(open_idx, len(s)):
        if s[i] == "(":
            depth += 1
        elif s[i] == ")":
            depth -= 1
            if depth == 0:
                return i
    return -1


_RE_SUMIFS = re.compile(r"SUMIFS\s*\(", re.IGNORECASE)
# célula de data: critério emparelhado com a coluna $B:$B de SERIE_VALOR
_RE_DATE_B = re.compile(r"\$?B\s*:\s*\$?B\s*,\s*(\$?[A-Z]{1,3}\$?\d+)", re.IGNORECASE)
_RE_CELL = re.compile(r"\$?[A-Z]{1,3}\$?\d+")


def _extrair_data_cell(args: str) -> str | None:
    """Extrai a célula de data de dentro dos argumentos do SUMIFS."""
    m = _RE_DATE_B.search(args)
    if m:
        return m.group(1)
    # fallback: última referência de célula (após remover a série numérica)
    cells = _RE_CELL.findall(args)
    return cells[-1] if cells else None


def sumifs_para_xlookup(formula: str) -> tuple[str, int]:
    """Substitui toda chamada SUMIFS(...) por XLOOKUP(<data>,CDI!A:A,CDI!B:B).

    Retorna (formula_nova, n_substituicoes).
    """
    out = formula
    n = 0
    while True:
        m = _RE_SUMIFS.search(out)
        if not m:
            break
        open_idx = m.end() - 1
        close = _match_paren(out, open_idx)
        if close == -1:
            break
        args = out[open_idx + 1:close]
        data_cell = _extrair_data_cell(args)
        if not data_cell:
            break
        # 4º arg (if_not_found = 0) reproduz o SUMIFS: datas fora do CDI (futuras)
        # devolvem 0 em vez de #N/A, evitando propagação de erro para PU/juros.
        repl = f"_xlfn.XLOOKUP({data_cell},CDI!A:A,CDI!B:B,0)"
        out = out[:m.start()] + repl + out[close + 1:]
        n += 1
    return out, n


# ---------------------------------------------------------------------------
# Localização de colunas
# ---------------------------------------------------------------------------
def _achar_taxa_data(ws):
    """Retorna (header_row, taxa_col, data_col) ou (0, None, None)."""
    for ri in range(1, 5):
        hdr = {str(c.value).strip(): c.column for c in ws[ri] if c.value is not None}
        if len(hdr) < 3:
            continue
        taxa = next((hdr[k] for k in NOMES_TAXA if k in hdr), None)
        if not taxa:
            continue
        data = next((hdr[k] for k in NOMES_DATA if k in hdr), None)
        return ri, taxa, data
    return 0, None, None


def _aba_calc(wb):
    for ws in wb.worksheets:
        if ws.title in ABAS_IGNORAR:
            continue
        hrow, taxa, data = _achar_taxa_data(ws)
        if taxa:
            return ws, hrow, taxa, data
    return None, 0, None, None


# ---------------------------------------------------------------------------
# Índice de arquivos homolog por nome de arquivo (robusto a nome de pasta)
# ---------------------------------------------------------------------------
def indexar_homolog() -> dict:
    idx = {}
    for sub in HOMO_ROOT.iterdir():
        if not sub.is_dir():
            continue
        for pasta in sub.iterdir():
            if not pasta.is_dir():
                continue
            for arq in pasta.iterdir():
                if arq.suffix.lower() in (".xlsx", ".xlsm"):
                    idx[arq.name.lower()] = arq
    return idx


# ---------------------------------------------------------------------------
# Correção de um arquivo
# ---------------------------------------------------------------------------
def corrigir_arquivo(sp_path: Path, homo_path: Path, tmp_dir: Path) -> dict:
    """Reescreve a coluna de taxa do homolog espelhando o SharePoint (SUMIFS->XLOOKUP)."""
    res = {"substituidos": 0, "verbatim": 0, "linhas": 0, "erro": None}

    # copia SP para tmp e lê (nunca abre o original diretamente)
    tmp_sp = tmp_dir / sp_path.name
    shutil.copy2(str(sp_path), str(tmp_sp))
    wb_sp = openpyxl.load_workbook(str(tmp_sp), data_only=False)
    ws_sp, hrow_sp, taxa_sp, data_sp = _aba_calc(wb_sp)
    if not ws_sp or not taxa_sp:
        wb_sp.close()
        res["erro"] = "SP sem coluna de taxa"
        return res

    wb_h = openpyxl.load_workbook(str(homo_path), data_only=False)
    ws_h, hrow_h, taxa_h, data_h = _aba_calc(wb_h)
    if not ws_h or not taxa_h:
        wb_sp.close(); wb_h.close()
        res["erro"] = "homolog sem coluna de taxa"
        return res

    if (ws_sp.title, taxa_sp, data_sp) != (ws_h.title, taxa_h, data_h):
        wb_sp.close(); wb_h.close()
        res["erro"] = (f"layout divergente SP({ws_sp.title},{taxa_sp},{data_sp}) "
                       f"!= HOMO({ws_h.title},{taxa_h},{data_h})")
        return res

    # última linha de dados = última com data na coluna de data (no homolog)
    ult = hrow_h
    for r in range(ws_h.max_row, hrow_h, -1):
        if ws_h.cell(r, data_h).value is not None:
            ult = r
            break

    for r in range(hrow_h + 1, ult + 1):
        sp_val = ws_sp.cell(r, taxa_sp).value
        if sp_val is None:
            continue
        res["linhas"] += 1
        if isinstance(sp_val, str) and "SUMIFS" in sp_val.upper():
            nova, n = sumifs_para_xlookup(sp_val)
            ws_h.cell(r, taxa_h).value = nova
            res["substituidos"] += 1
        else:
            # valor fixo ou fórmula sem SUMIFS -> copia verbatim do SP
            ws_h.cell(r, taxa_h).value = sp_val
            res["verbatim"] += 1

    wb_sp.close()
    wb_h.save(str(homo_path))
    wb_h.close()
    return res


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------
def main(dry_run: bool, todos: bool) -> None:
    rows = ler_dict(str(config.CADASTRO))
    ativos = {
        r["IdSerie"]: r
        for r in rows
        if r.get("CalcPath")
        and r.get("Familia", "") in FAMILIAS_CDI
        and r.get("Status", "").lower() != "liquidado"
    }

    alvos = list(ativos.keys()) if todos else IDS_CORRIGIR
    idx = indexar_homolog()

    print(f"Corrigindo {len(alvos)} calculadora(s)")
    print(f"Homolog root: {HOMO_ROOT}")
    print("(DRY RUN)\n" if dry_run else "")

    tmp = Path(tempfile.mkdtemp(prefix="corr_procx_"))
    ok = erros = 0
    try:
        for sid in alvos:
            r = ativos.get(sid)
            if not r:
                print(f"  [{sid:>10}] NAO NO CADASTRO")
                erros += 1
                continue
            ape = r.get("Apelido", sid)
            sp = Path(config.FLUXO_DIR) / r["CalcPath"]
            fname = Path(r["CalcPath"]).name
            homo = idx.get(fname.lower())

            if not sp.exists():
                print(f"  [{sid:>10}] {ape:<24} SP NAO EXISTE: {fname}")
                erros += 1
                continue
            if not homo:
                print(f"  [{sid:>10}] {ape:<24} HOMOLOG NAO ACHADO: {fname}")
                erros += 1
                continue

            if dry_run:
                # só mostra o que faria: conta SUMIFS no SP
                tmp_sp = tmp / sp.name
                shutil.copy2(str(sp), str(tmp_sp))
                wb = openpyxl.load_workbook(str(tmp_sp), data_only=False, read_only=True)
                # localizar taxa
                n_sumifs = 0
                exemplo = ""
                for ws in wb.worksheets:
                    if ws.title in ABAS_IGNORAR:
                        continue
                    rows_c = list(ws.iter_rows(values_only=False))
                    hdr = {}; hrow = 0
                    for ri in range(min(4, len(rows_c))):
                        v = {str(c.value).strip(): c.column for c in rows_c[ri] if c.value is not None}
                        if len(v) >= 3:
                            hdr = v; hrow = ri + 1; break
                    tcol = next((hdr[k] for k in NOMES_TAXA if k in hdr), None)
                    if not tcol:
                        continue
                    for ri in range(hrow, len(rows_c)):
                        val = rows_c[ri][tcol - 1].value if tcol - 1 < len(rows_c[ri]) else None
                        if isinstance(val, str) and "SUMIFS" in val.upper():
                            n_sumifs += 1
                            if not exemplo:
                                exemplo = sumifs_para_xlookup(val)[0]
                    break
                wb.close()
                print(f"  [{sid:>10}] {ape:<24} {n_sumifs} SUMIFS -> XLOOKUP")
                if exemplo:
                    print(f"               ex: {exemplo[:80]}")
                ok += 1
            else:
                res = corrigir_arquivo(sp, homo, tmp)
                if res["erro"]:
                    print(f"  [{sid:>10}] {ape:<24} ERRO: {res['erro']}")
                    erros += 1
                else:
                    print(f"  [{sid:>10}] {ape:<24} "
                          f"{res['substituidos']} XLOOKUP + {res['verbatim']} fixos "
                          f"({res['linhas']} linhas)")
                    ok += 1
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    print(f"\n{'Simuladas' if dry_run else 'Corrigidas'}: {ok} | Erros: {erros}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    main(dry_run="--dry-run" in sys.argv, todos="--todos" in sys.argv)

"""Adiciona o valor-padrão (if_not_found = 0) às fórmulas XLOOKUP das calculadoras.

Defeito corrigido: a migração SUMIFS→XLOOKUP não incluiu o 4º argumento do
XLOOKUP. Para datas fora do CDI (futuras, além da última data de mercado) o
XLOOKUP retornava `#N/A`, que propagava para PU/juros e "corrompia" a planilha.
O SUMIFS original retornava 0 nesses casos (accrual futuro zero). A correção
reproduz esse comportamento:

    =XLOOKUP(data, CDI!A:A, CDI!B:B)   →   =XLOOKUP(data, CDI!A:A, CDI!B:B, 0)

Escrita em subprocessos de poucos arquivos (a proteção antiransomware do TI mata
processos que sobrescrevem muitos documentos Office em sequência).

Uso:
    python -m mtm_skills.corrigir_xlookup_default            # todas
    python -m mtm_skills.corrigir_xlookup_default --check    # só verifica
    python -m mtm_skills.corrigir_xlookup_default <arq...>   # (modo subprocesso)
"""
import re
import subprocess
import sys
from pathlib import Path

_RAIZ = Path(__file__).resolve().parents[1]
if str(_RAIZ) not in sys.path:
    sys.path.insert(0, str(_RAIZ))

import openpyxl

from pu_mtm.app import config

BATCH = 3
# Fecha um XLOOKUP(...,CDI!B:B) SEM o 4º argumento. Idempotente: depois vira
# "CDI!B:B,0)" e o padrão deixa de casar.
_RE_SEM_DEFAULT = re.compile(r"CDI!B:B\s*\)")


def _fix_um(path: Path) -> int:
    wb = openpyxl.load_workbook(str(path), data_only=False)
    n = 0
    for ws in wb.worksheets:
        if ws.title in ("CDI", "Feriados"):
            continue
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "XLOOKUP" in v.upper() and _RE_SEM_DEFAULT.search(v):
                    cell.value = _RE_SEM_DEFAULT.sub("CDI!B:B,0)", v)
                    n += 1
    if n:
        wb.save(str(path))
    wb.close()
    return n


def _contar_pendentes(path: Path) -> int:
    wb = openpyxl.load_workbook(str(path), data_only=False, read_only=True)
    n = 0
    for ws in wb.worksheets:
        if ws.title in ("CDI", "Feriados"):
            continue
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and "XLOOKUP" in v.upper() and _RE_SEM_DEFAULT.search(v):
                    n += 1
    wb.close()
    return n


def _arquivos() -> list[Path]:
    return sorted(
        p for p in config.HOMOLOG_CALC_ROOT.rglob("*")
        if p.suffix.lower() in (".xlsx", ".xlsm")
    )


def check() -> None:
    total = 0
    for p in _arquivos():
        pend = _contar_pendentes(p)
        total += pend
        marca = f"  <<< {pend} sem default" if pend else ""
        print(f"  {pend:>4} pendente(s)  {p.parent.name}{marca}")
    print(f"\nTotal de XLOOKUP sem valor-padrão: {total}")


def aplicar() -> None:
    arqs = _arquivos()
    lotes = [arqs[i:i + BATCH] for i in range(0, len(arqs), BATCH)]
    print(f"Corrigindo {len(arqs)} arquivo(s) em {len(lotes)} subprocesso(s) de até {BATCH}...\n")
    for lote in lotes:
        proc = subprocess.run(
            [sys.executable, "-W", "ignore", str(Path(__file__).resolve()),
             *[str(p) for p in lote]],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
        )
        if proc.stdout:
            sys.stdout.write(proc.stdout)
        if proc.returncode != 0 and proc.stderr:
            sys.stdout.write(proc.stderr)
    print("\nReverificando...")
    check()


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if "--check" in sys.argv:
        check()
    elif args:
        # modo subprocesso: corrige os arquivos passados e sai
        for a in args:
            p = Path(a)
            n = _fix_um(p)
            print(f"  {p.parent.name}: +{n} XLOOKUP com default")
    else:
        aplicar()

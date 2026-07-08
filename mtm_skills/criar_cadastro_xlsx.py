"""Gera (ou regera) data/cadastro_ativos.xlsx a partir do cadastro_ativos.csv.

Estrutura da planilha gerada:
  - "Instruções"  : guia de preenchimento com tabela de colunas e exemplos
  - "Ativos"      : dados migrados + dropdowns de validação
  - "Listas"      : aba oculta com as listas usadas pelos dropdowns

A planilha xlsx passa a ser a fonte única do cadastro; o CSV original é mantido
como backup. Após rodar este script uma vez, edite apenas o xlsx.

Uso:
    python -m mtm_skills.criar_cadastro_xlsx
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_RAIZ = Path(__file__).resolve().parents[1]
if str(_RAIZ) not in sys.path:
    sys.path.insert(0, str(_RAIZ))

from pu_mtm.app import config
from pu_mtm.dados import csvio

# ---------------------------------------------------------------------------
# Colunas (mesma ordem do CSV — NÃO alterar sem atualizar o código leitor)
# ---------------------------------------------------------------------------
_COLS = [
    "IdSerie", "IdTitulo", "Apelido", "Familia", "Indexador", "SerieIndice",
    "TaxaFixa", "VNe", "Spread", "DataEmissao", "CalcPath", "AbaPU",
    "CelulaPU", "Status", "Base", "FatorDiarioArred", "JurosArred",
    "EvJurosFonte", "EvAmortFonte", "DataCol", "PercentualCDI", "PreDC",
    "PreDaycount",
]

# ---------------------------------------------------------------------------
# Conversão Familia: código interno → nome amigável (e vice-versa em csvio)
# ---------------------------------------------------------------------------
_INT_TO_FRIENDLY = {
    "di_puro":    "DI",
    "di_spread":  "DI + SPREAD",
    "prefixado":  "PREFIXADO",
    "ipca_spread": "IPCA + SPREAD",
}

# ---------------------------------------------------------------------------
# Listas de validação (coluna na aba Listas → valores)
# ---------------------------------------------------------------------------
_LISTAS = [
    ("Familia",         ["DI", "DI + SPREAD", "PREFIXADO", "IPCA + SPREAD", "IPCA"]),
    ("Indexador",       ["DI", "PRE", "IPCA"]),
    ("Status",          ["verificado", "liquidado", "congelado"]),
    ("Base",            [252, 360, 365, 30]),
    ("JurosArred",      ["trunc8", "nenhum", "round8"]),
    ("FatorDiarioArred",["nenhum", "round8", "cru"]),
    ("EvJurosFonte",    ["agendado", "nenhum"]),
    ("EvAmortFonte",    ["agendado", "nenhum"]),
    ("PreDaycount",     ["30360", "actual", "mensal"]),
]

# coluna (nome) → fórmula de referência na aba Listas
def _lista_formula(col_nome: str) -> str | None:
    for i, (nome, vals) in enumerate(_LISTAS, 1):
        if nome == col_nome:
            letra = get_column_letter(i)
            return f"Listas!${letra}$2:${letra}${1 + len(vals)}"
    return None

# ---------------------------------------------------------------------------
# Guia de colunas para aba Instruções
# ---------------------------------------------------------------------------
_GUIDE = [
    # (Campo, Obrigatório, Quando usar, Padrão, Descrição)
    ("IdSerie", "Sim", "Sempre",
     "—",
     "ID único do ativo no sistema. Número inteiro; não pode se repetir."),
    ("IdTitulo", "Sim", "Sempre",
     "—",
     "Código do título na OSLO. Pode ser alfanumérico (ex: 24B01290474)."),
    ("Apelido", "Sim", "Sempre",
     "—",
     "Nome curto do ativo para identificação nos relatórios."),
    ("Familia", "Sim", "Sempre",
     "—",
     "Família de precificação — determina o modelo matemático:\n"
     "DI = accrual CDI diário (acordo, mútuo)\n"
     "DI + SPREAD = CDI acumulado + spread\n"
     "PREFIXADO = taxa fixa (curva prefixada)\n"
     "IPCA + SPREAD = VNA NTN-B + spread sobre IPCA"),
    ("Indexador", "Sim", "Sempre",
     "—",
     "Índice de referência: DI (CDI diário), PRE (prefixado), IPCA."),
    ("SerieIndice", "DI e IPCA", "Preencher quando Indexador = DI ou IPCA",
     "—",
     "Série do índice na base de mercado. CDI = 2, IPCA = 1."),
    ("TaxaFixa", "PREFIXADO", "Cupom ou yield anual",
     "0",
     "Taxa fixa anual em decimal. Exemplo: 2% a.a. → 0.02"),
    ("VNe", "Sim", "Sempre",
     "—",
     "Valor Nominal de Emissão em R$. NCs e similares usam VNe = 1. "
     "Acordos usam o valor nominal inteiro (ex: 750000)."),
    ("Spread", "DI+SPREAD e IPCA+SPREAD", "Quando Familia tem spread",
     "0",
     "Spread anual sobre o índice, em decimal. Exemplo: 5% a.a. → 0.05"),
    ("DataEmissao", "Sim", "Sempre",
     "—",
     "Data de emissão ou data-âncora da calculadora. Formato DD/MM/AAAA."),
    ("CalcPath", "Sim", "Sempre",
     "—",
     "Caminho RELATIVO da calculadora a partir da pasta 'Fluxo de pagamento'. "
     "Use / ou \\ como separador.\n"
     "Exemplo: Contratos de Mútuo.../calculadora.xlsx"),
    ("AbaPU", "Sim", "Sempre",
     "Sheet1",
     "Nome exato da aba na calculadora que contém o histórico de PU."),
    ("CelulaPU", "Sim", "Sempre",
     "M",
     "Letra da coluna onde está o PU na aba indicada em AbaPU. "
     "Exemplos: M, N, K, P."),
    ("Status", "Sim", "Sempre",
     "verificado",
     "Status do ativo:\n"
     "verificado = ativo, calculando normalmente\n"
     "liquidado  = encerrado (ignorado pelo motor)\n"
     "congelado  = PU fixo, sem accrual"),
    ("Base", "DI+SPREAD e PREFIXADO", "Base de capitalização dos juros",
     "252",
     "Base de dias para capitalização: 252 (dias úteis), 360 ou 365 (corridos), "
     "30 (mensal/comercial)."),
    ("FatorDiarioArred", "Técnico", "Raramente necessário",
     "nenhum",
     "Arredondamento do fator DI diário:\n"
     "nenhum = sem arredondamento (padrão)\n"
     "round8 = arredonda em 8 casas decimais\n"
     "cru    = produto bruto sem qualquer arredondamento"),
    ("JurosArred", "Técnico", "Raramente necessário",
     "trunc8",
     "Arredondamento dos juros acumulados:\n"
     "trunc8 = trunca em 8 casas decimais (padrão)\n"
     "nenhum = sem arredondamento\n"
     "round8 = arredonda em 8 casas"),
    ("EvJurosFonte", "Ativos com eventos", "Quando há pagamentos de juros agendados",
     "nenhum",
     "Como eventos de juros são lidos:\n"
     "agendado = lê do arquivo data/eventos/{IdSerie}.xlsx\n"
     "nenhum   = ativo sem eventos de juros"),
    ("EvAmortFonte", "Ativos com eventos", "Quando há amortizações agendadas",
     "nenhum",
     "Como eventos de amortização são lidos:\n"
     "agendado = lê do arquivo de eventos\n"
     "nenhum   = ativo sem amortização"),
    ("DataCol", "Técnico", "Raramente necessário",
     "C",
     "Letra da coluna com as datas na calculadora. Padrão = C."),
    ("PercentualCDI", "DI com % ≠ 100%", "Quando remuneração é X% do CDI",
     "100",
     "Percentual do CDI. Exemplo: ativo a 105% do CDI → preencher 105."),
    ("PreDC", "PREFIXADO", "Convenção de dias corridos",
     "30",
     "Base de dias corridos do prefixado: 30, 252, 360 ou 365."),
    ("PreDaycount", "PREFIXADO", "Convenção de contagem de dias",
     "30360",
     "Convenção do prefixado:\n"
     "30360   = convenção bancária (mês 30 dias, ano 360)\n"
     "actual  = dias corridos reais\n"
     "mensal  = capitalização mensal"),
]


# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border_thin() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


_HDR_FILL   = _fill("1F4E79")     # cabeçalho Ativos: azul escuro
_HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
_GUIDE_HDR  = _fill("2E75B6")     # cabeçalho guia: azul médio
_GUIDE_FONT = Font(bold=True, color="FFFFFF", size=10)
_TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
_SEC_FONT   = Font(bold=True, size=11, color="2E75B6")
_BODY_FONT  = Font(size=10)
_WRAP       = Alignment(wrap_text=True, vertical="top")
_CENTER     = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Aba Listas (oculta)
# ---------------------------------------------------------------------------
def _criar_listas(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    ws = wb.create_sheet("Listas")
    for col_idx, (nome, vals) in enumerate(_LISTAS, 1):
        letra = get_column_letter(col_idx)
        ws[f"{letra}1"] = nome
        ws[f"{letra}1"].font = Font(bold=True)
        for row_idx, v in enumerate(vals, 2):
            ws[f"{letra}{row_idx}"] = v
    ws.sheet_state = "hidden"
    return ws


# ---------------------------------------------------------------------------
# Aba Instruções
# ---------------------------------------------------------------------------
def _criar_instrucoes(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    ws = wb.create_sheet("Instruções")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 60

    r = 1

    # Título
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = "CADASTRO DE ATIVOS — GUIA DE PREENCHIMENTO"
    ws[f"A{r}"].font = _TITLE_FONT
    ws[f"A{r}"].alignment = _CENTER
    ws.row_dimensions[r].height = 28
    r += 1

    ws[f"A{r}"] = (
        "Esta planilha é a fonte de dados do motor de precificação MtM. "
        "Cada linha na aba 'Ativos' representa um ativo da carteira."
    )
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"].font = _BODY_FONT
    ws[f"A{r}"].alignment = _WRAP
    ws.row_dimensions[r].height = 30
    r += 1

    r += 1  # linha vazia

    # Avisos
    ws[f"A{r}"] = "⚠  REGRAS IMPORTANTES"
    ws[f"A{r}"].font = _SEC_FONT
    ws.merge_cells(f"A{r}:E{r}")
    r += 1

    avisos = [
        "Não altere os nomes das colunas (primeira linha da aba 'Ativos').",
        "Não reordene as colunas.",
        "IdSerie deve ser único — nunca repita o mesmo número.",
        "Nunca delete linhas de ativos liquidados: mude o Status para 'liquidado'.",
        "CalcPath é relativo à pasta 'Fluxo de pagamento' — não use caminho absoluto.",
    ]
    for av in avisos:
        ws[f"A{r}"] = f"  •  {av}"
        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"].font = _BODY_FONT
        ws[f"A{r}"].alignment = _WRAP
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1

    # Como cadastrar
    ws[f"A{r}"] = "COMO CADASTRAR UM NOVO ATIVO"
    ws[f"A{r}"].font = _SEC_FONT
    ws.merge_cells(f"A{r}:E{r}")
    r += 1

    passos = [
        "1. Vá para a aba 'Ativos'.",
        "2. Adicione uma nova linha ao final da tabela.",
        "3. Preencha todas as colunas obrigatórias (ver tabela abaixo).",
        "4. Use os menus suspensos (dropdown) para Familia, Indexador, Status e colunas técnicas.",
        "5. Salve e feche a planilha antes de rodar o motor.",
    ]
    for p in passos:
        ws[f"A{r}"] = f"  {p}"
        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"].font = _BODY_FONT
        ws[f"A{r}"].alignment = _WRAP
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1

    # Tabela de colunas
    ws[f"A{r}"] = "GUIA DE COLUNAS"
    ws[f"A{r}"].font = _SEC_FONT
    ws.merge_cells(f"A{r}:E{r}")
    r += 1

    hdrs = ["Campo", "Obrigatório", "Quando usar", "Padrão", "Descrição"]
    for col_idx, h in enumerate(hdrs, 1):
        c = ws.cell(row=r, column=col_idx, value=h)
        c.font = _GUIDE_FONT
        c.fill = _GUIDE_HDR
        c.alignment = _CENTER
        c.border = _border_thin()
    ws.row_dimensions[r].height = 20
    r += 1

    for campo, obrig, quando, padrao, desc in _GUIDE:
        vals = [campo, obrig, quando, padrao, desc]
        for col_idx, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col_idx, value=v)
            c.font = _BODY_FONT
            c.alignment = _WRAP
            c.border = _border_thin()
        ws.row_dimensions[r].height = max(18, desc.count("\n") * 15 + 18)
        r += 1

    r += 1

    # Exemplos por família
    ws[f"A{r}"] = "EXEMPLOS POR FAMÍLIA"
    ws[f"A{r}"].font = _SEC_FONT
    ws.merge_cells(f"A{r}:E{r}")
    r += 1

    exemplos = [
        ("DI (accrual CDI — acordos, mútuo)",
         "IdSerie: 531 | Familia: DI | Indexador: DI | SerieIndice: 2 | VNe: 750000 | DataEmissao: 25/06/2024\n"
         "Colunas técnicas (Base, JurosArred...): deixar em branco → motor usa padrão."),
        ("DI + SPREAD (CDI + spread — NCs, Sport Media)",
         "IdSerie: 741 | Familia: DI + SPREAD | Indexador: DI | SerieIndice: 2 | Spread: 0.05 | VNe: 1000\n"
         "Base: 252 | JurosArred: trunc8"),
        ("PREFIXADO (taxa fixa — NCs, CCBs)",
         "IdSerie: 476 | Familia: PREFIXADO | Indexador: PRE | TaxaFixa: 0.02 | VNe: 1\n"
         "Base: 360 | PreDC: 30 | PreDaycount: 30360"),
        ("IPCA + SPREAD (VNA NTN-B + spread)",
         "IdSerie: 367 | Familia: IPCA + SPREAD | Indexador: IPCA | SerieIndice: 1 | Spread: 0.15 | VNe: 1"),
    ]

    for titulo, detalhe in exemplos:
        ws[f"A{r}"] = f"  {titulo}"
        ws[f"A{r}"].font = Font(bold=True, size=10)
        ws.merge_cells(f"A{r}:E{r}")
        ws.row_dimensions[r].height = 18
        r += 1
        ws[f"A{r}"] = f"    {detalhe}"
        ws[f"A{r}"].font = _BODY_FONT
        ws[f"A{r}"].alignment = _WRAP
        ws.merge_cells(f"A{r}:E{r}")
        ws.row_dimensions[r].height = 36
        r += 1

    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------------------
# Aba Ativos
# ---------------------------------------------------------------------------
def _criar_ativos(
    wb: openpyxl.Workbook, rows: list[dict]
) -> openpyxl.worksheet.worksheet.Worksheet:
    ws = wb.create_sheet("Ativos")

    # Cabeçalho
    for col_idx, col_name in enumerate(_COLS, 1):
        c = ws.cell(row=1, column=col_idx, value=col_name)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = _CENTER
        c.border = _border_thin()

    # Colunas que devem ser armazenadas como texto (format "@") mesmo sendo numéricas
    _TEXT_COLS = {"IdTitulo"}

    # Dados migrados
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(_COLS, 1):
            raw = row.get(col_name, "")
            # Familia: código interno → nome amigável
            if col_name == "Familia":
                val = _INT_TO_FRIENDLY.get(raw, raw)
            # Colunas forçadas como texto
            elif col_name in _TEXT_COLS:
                val = str(raw) if raw else ""
            # Números inteiros
            elif col_name in ("IdSerie", "SerieIndice", "Base", "PercentualCDI", "PreDC"):
                try:
                    val = int(float(raw)) if raw else ""
                except (ValueError, TypeError):
                    val = raw
            # Números decimais
            elif col_name in ("TaxaFixa", "VNe", "Spread"):
                try:
                    val = float(raw) if raw else ""
                except (ValueError, TypeError):
                    val = raw
            else:
                val = raw
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = _border_thin()
            c.font = Font(size=10)
            if col_name in _TEXT_COLS:
                c.number_format = "@"  # força texto; evita Excel auto-converter "569190" → número

    # Aplica formato texto também nas linhas futuras (vazias) da coluna IdTitulo
    max_row = len(rows) + 500
    for col_idx, col_name in enumerate(_COLS, 1):
        if col_name in _TEXT_COLS:
            for row_idx in range(len(rows) + 2, max_row + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.number_format = "@"

    # Larguras aproximadas
    _widths = {
        "IdSerie": 10, "IdTitulo": 14, "Apelido": 28, "Familia": 16,
        "Indexador": 12, "SerieIndice": 13, "TaxaFixa": 12, "VNe": 14,
        "Spread": 10, "DataEmissao": 14, "CalcPath": 55, "AbaPU": 18,
        "CelulaPU": 12, "Status": 14, "Base": 8, "FatorDiarioArred": 18,
        "JurosArred": 12, "EvJurosFonte": 14, "EvAmortFonte": 14,
        "DataCol": 10, "PercentualCDI": 14, "PreDC": 10, "PreDaycount": 14,
    }
    for col_idx, col_name in enumerate(_COLS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _widths.get(col_name, 14)

    # Validações por dropdown
    max_row = max(len(rows) + 500, 200)
    for col_idx, col_name in enumerate(_COLS, 1):
        formula = _lista_formula(col_name)
        if formula is None:
            continue
        letra = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Valor inválido",
            error=f"Selecione um valor da lista para '{col_name}'.",
        )
        dv.sqref = f"{letra}2:{letra}{max_row}"
        ws.add_data_validation(dv)

    # Painel congelado + filtro
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLS))}1"

    return ws


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------
def criar_xlsx(csv_path: str | None = None, xlsx_path: str | None = None) -> str:
    xlsx_path = xlsx_path or str(config.DATA / "cadastro_ativos.xlsx")

    # Fonte dos dados: xlsx existente (re-formatação) ou csv (migração inicial)
    if csv_path:
        source = csv_path
    elif Path(xlsx_path).exists():
        source = xlsx_path   # xlsx já é fonte de verdade; relê dele mesmo
    else:
        source = str(config.DATA / "cadastro_ativos.csv")

    try:
        rows = csvio.ler_dict(source)
    except PermissionError:
        raise SystemExit(
            f"\nArquivo aberto no Excel — feche-o antes de rodar este script:\n  {source}"
        )
    print(f"Lidos {len(rows)} ativos de: {source}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _criar_listas(wb)        # criada 1ª (vai ficar oculta, mas precisa existir antes da validação)
    _criar_instrucoes(wb)
    _criar_ativos(wb, rows)

    # Ordem das abas: Instruções | Ativos | Listas(hidden)
    wb.move_sheet("Instruções", offset=-(len(wb.sheetnames) - 1))
    wb.move_sheet("Ativos",     offset=-(len(wb.sheetnames) - 2))

    wb.save(xlsx_path)
    print(f"Planilha gerada: {xlsx_path}")
    return xlsx_path


if __name__ == "__main__":
    criar_xlsx()

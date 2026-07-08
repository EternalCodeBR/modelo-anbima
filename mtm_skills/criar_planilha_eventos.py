"""Cria (ou atualiza) planilhas de eventos estilizadas para cada ativo do cadastro.

Para cada ativo:
  - Lê o CSV legado se existir e migra os dados
  - Gera data/Eventos/{FAMILIA}/{id_serie}/{id_serie}.xlsx com:
      * Aba "Eventos": tabela Previsto/Realizado, Tipo (dropdown), Observacao
      * Aba "Info": metadados do ativo
  - Se o xlsx já existe, pula (não sobrescreve dados reais)

Uso:
    python -m mtm_skills.criar_planilha_eventos
    python -m mtm_skills.criar_planilha_eventos --forcar
"""
import argparse
import sys
from pathlib import Path

_RAIZ = Path(__file__).resolve().parents[1]
if str(_RAIZ) not in sys.path:
    sys.path.insert(0, str(_RAIZ))

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from pu_mtm.app import config
from pu_mtm.dados import csvio
from pu_mtm.dados.cadastro import ler_cadastro

_FAMILIA_PASTA = {
    "di_puro":    "CDI",
    "di_spread":  "CDI + SPREAD",
    "prefixado":  "PREFIXADO",
    "ipca_spread": "IPCA",
}

# ---- Paleta ----
_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # navy
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_INFO_FILL    = PatternFill("solid", fgColor="D9E1F2")   # azul claro (rótulo Info)
_PAGO_FILL    = PatternFill("solid", fgColor="E2EFDA")   # verde (pago)
_EXTRA_FILL   = PatternFill("solid", fgColor="DEEBF7")   # azul (extraordinário)
_PARC_FILL    = PatternFill("solid", fgColor="FCE4D6")   # laranja (parcial)
_ATRASO_FILL  = PatternFill("solid", fgColor="FFF2CC")   # amarelo (atrasado)
_ANTEC_FILL   = PatternFill("solid", fgColor="EAF7EA")   # verde suave (antecipado)

_THIN   = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_COLUNAS = [
    ("DataPrevista",          14),
    ("DataRealizada",         14),
    ("JurosPrevisto",         18),
    ("JurosRealizado",        18),
    ("AmortizacaoPrevista",   22),
    ("AmortizacaoRealizada",  22),
    ("Tipo",                  15),
    ("Observacao",            38),
]

_TIPOS_VALIDOS = "agendado,extraordinario,parcial,antecipado,atrasado"
_FMT_NUM = "0.00000000"


# --------------------------------------------------------------------------- #
#  Helpers
# --------------------------------------------------------------------------- #
def _num(v: str) -> float | str:
    try:
        return float(v) if v not in (None, "") else ""
    except (ValueError, TypeError):
        return v if v else ""


def _fill_linha(tipo: str, data_realizada: str):
    if data_realizada:
        return _PAGO_FILL
    match tipo:
        case "extraordinario": return _EXTRA_FILL
        case "parcial":        return _PARC_FILL
        case "atrasado":       return _ATRASO_FILL
        case "antecipado":     return _ANTEC_FILL
        case _:                return None


def _ler_csv_legado(caminho: Path) -> list[dict]:
    """Converte CSV antigo (Data/EventoJuros/EventoAmortizacao/Obs) para o formato novo."""
    rows = []
    for r in csvio.ler_dict(str(caminho)):
        rows.append({
            "DataPrevista":         r.get("Data", ""),
            "DataRealizada":        "",
            "JurosPrevisto":        r.get("EventoJuros", ""),
            "JurosRealizado":       "",
            "AmortizacaoPrevista":  r.get("EventoAmortizacao", ""),
            "AmortizacaoRealizada": "",
            "Tipo":                 "agendado",
            "Observacao":           r.get("Obs", ""),
        })
    return rows


# --------------------------------------------------------------------------- #
#  Aba Eventos
# --------------------------------------------------------------------------- #
def _aba_eventos(wb, dados: list[dict]):
    ws = wb.create_sheet("Eventos")

    # --- Cabeçalho ---
    for col, (nome, largura) in enumerate(_COLUNAS, 1):
        cell = ws.cell(1, col, nome)
        cell.font   = _HEADER_FONT
        cell.fill   = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = largura
    ws.row_dimensions[1].height = 22

    # --- Linhas de dados ---
    for d in dados:
        ws.append([
            d.get("DataPrevista", ""),
            d.get("DataRealizada", ""),
            _num(d.get("JurosPrevisto", "")),
            _num(d.get("JurosRealizado", "")),
            _num(d.get("AmortizacaoPrevista", "")),
            _num(d.get("AmortizacaoRealizada", "")),
            d.get("Tipo", "agendado"),
            d.get("Observacao", ""),
        ])
        row = ws.max_row
        fill = _fill_linha(d.get("Tipo", "agendado"), d.get("DataRealizada", ""))
        for col in range(1, len(_COLUNAS) + 1):
            cell = ws.cell(row, col)
            cell.border = _BORDER
            if fill:
                cell.fill = fill
            if col in (3, 4, 5, 6):          # numéricas
                cell.number_format = _FMT_NUM
                cell.alignment = Alignment(horizontal="right")
            elif col in (1, 2, 7):            # datas + tipo
                cell.alignment = Alignment(horizontal="center")

    # --- Dropdown Tipo ---
    n_max = max(ws.max_row + 50, 200)
    dv = DataValidation(
        type="list",
        formula1=f'"{_TIPOS_VALIDOS}"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.sqref = f"G2:G{n_max}"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUNAS))}1"
    return ws


# --------------------------------------------------------------------------- #
#  Aba Info
# --------------------------------------------------------------------------- #
def _aba_info(wb, id_serie: str, ativo):
    ws = wb.create_sheet("Info")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55

    campos = [
        ("IdSerie",     id_serie),
        ("Apelido",     ativo.apelido),
        ("Família",     ativo.familia),
        ("VNe",         ativo.vne),
        ("DataEmissao", ativo.data_emissao.strftime("%d/%m/%Y")),
        ("Spread",      ativo.spread),
        ("TaxaFixa",    ativo.taxa_fixa),
        ("Base",        ativo.base),
    ]
    for label, val in campos:
        ws.append([label, str(val)])
        row = ws.max_row
        c_rot = ws.cell(row, 1)
        c_rot.font   = Font(bold=True)
        c_rot.fill   = _INFO_FILL
        c_rot.border = _BORDER
        ws.cell(row, 2).border = _BORDER

    # Legenda de tipos
    ws.append([])
    ws.append(["Tipos de Evento"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color="1F3864")

    legenda = [
        ("agendado",       "Evento programado no contrato"),
        ("extraordinario", "Amortização ou juros fora do calendário original"),
        ("parcial",        "Pagamento menor que o previsto (capitaliza a diferença)"),
        ("antecipado",     "Pagamento antes da data programada"),
        ("atrasado",       "Pagamento após a data programada — ativo fica congelado no PU do vencimento"),
    ]
    fills_leg = [None, _EXTRA_FILL, _PARC_FILL, _ANTEC_FILL, _ATRASO_FILL]
    for (tipo, desc), fill_l in zip(legenda, fills_leg):
        ws.append([tipo, desc])
        row = ws.max_row
        ws.cell(row, 1).border = _BORDER
        ws.cell(row, 2).border = _BORDER
        if fill_l:
            ws.cell(row, 1).fill = fill_l
            ws.cell(row, 2).fill = fill_l

    return ws


# --------------------------------------------------------------------------- #
#  Ponto de entrada por ativo / por todos
# --------------------------------------------------------------------------- #
def criar_planilha_ativo(id_serie: str, ativo, forcar: bool = False) -> Path | None:
    pasta_familia = _FAMILIA_PASTA.get(ativo.familia, "")
    if not pasta_familia:
        print(f"  Família desconhecida '{ativo.familia}' para {id_serie}, pulado.")
        return None

    pasta_ativo = config.EVENTOS_DIR / pasta_familia / id_serie
    pasta_ativo.mkdir(parents=True, exist_ok=True)
    xlsx_path = pasta_ativo / f"{id_serie}.xlsx"

    if xlsx_path.exists() and not forcar:
        return None

    csv_path = pasta_ativo / f"{id_serie}.csv"
    dados = _ler_csv_legado(csv_path) if csv_path.exists() else []

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _aba_eventos(wb, dados)
    _aba_info(wb, id_serie, ativo)

    wb.save(str(xlsx_path))
    return xlsx_path


def criar_todas(forcar: bool = False) -> list[Path]:
    ativos = ler_cadastro(str(config.CADASTRO))
    criados: list[Path] = []
    pulados: list[str] = []

    for id_serie, ativo in ativos.items():
        result = criar_planilha_ativo(id_serie, ativo, forcar=forcar)
        if result:
            criados.append(result)
        else:
            pulados.append(id_serie)

    print(f"\nPlanilhas criadas : {len(criados)}")
    for p in criados:
        try:
            rel = p.relative_to(_RAIZ)
        except ValueError:
            rel = p
        print(f"  + {rel}")
    if pulados:
        print(f"Pulados (já existem): {len(pulados)} — use --forcar para recriar")
    return criados


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Gera planilhas de eventos para todos os ativos")
    parser.add_argument("--forcar", action="store_true", help="Recria mesmo se xlsx já existir")
    args = parser.parse_args()
    criar_todas(forcar=args.forcar)

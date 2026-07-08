"""Cria/atualiza o workbook 'Rotina_Recalc' que hospeda a macro de recálculo.

Divisão de trabalho da rotina de CDI:
  1. Python (atualizar_cdi_surgical) grava o CDI no XML das calculadoras.
  2. Este workbook, via a macro VBA ``RecalcCache``, abre cada calculadora,
     recalcula (XLOOKUP local) e salva — criando o cache de PU do dia.

Este script cuida só da aba ``Lista`` (os caminhos que a macro vai abrir):
  * se ``Rotina_Recalc.xlsm`` já existe, atualiza a aba Lista PRESERVANDO a macro
    (openpyxl com keep_vba=True — o workbook é um runner simples, sem VML/desenho,
    então o round-trip do openpyxl aqui é seguro);
  * senão, cria ``Rotina_Recalc.xlsx`` base (você adiciona a macro no VBE — Alt+F11
    — e salva como .xlsm uma única vez).

Uso:
    python -m mtm_skills.criar_rotina_recalc
"""
from __future__ import annotations

import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

_RAIZ = Path(__file__).resolve().parents[1]
if str(_RAIZ) not in sys.path:
    sys.path.insert(0, str(_RAIZ))

import openpyxl
from openpyxl.styles import Font

from pu_mtm.app import config

HOMOLOG_ROOT = config.HOMOLOG_CALC_ROOT
XLSM = _RAIZ / "Rotina_Recalc.xlsm"
XLSX = _RAIZ / "Rotina_Recalc.xlsx"

INSTRUCOES = [
    "ROTINA DE RECÁLCULO DAS CALCULADORAS (cache de PU)",
    "",
    "Fluxo diário:",
    "  1. (Python) python -m mtm_skills.atualizar_cdi_surgical   -> grava o CDI novo no XML",
    "  2. (aqui)   rodar a macro RecalcCache                      -> recalcula e salva (cache)",
    "",
    "Configuração da macro (uma vez):",
    "  a. Abra o VBE com Alt+F11.",
    "  b. Inserir > Módulo. Cole o código da macro RecalcCache (fornecido à parte).",
    "  c. Salve este arquivo como 'Rotina_Recalc.xlsm' (Pasta de Trabalho Habilitada para Macro).",
    "",
    "Para rodar: Alt+F8 > RecalcCache > Executar (ou um botão nesta aba).",
    "",
    "A aba 'Lista' tem os caminhos que a macro abre. Para atualizá-la depois",
    "(ex.: mudou o conjunto de calculadoras): rode de novo",
    "  python -m mtm_skills.criar_rotina_recalc",
    "que reescreve só a Lista e preserva a macro.",
]


def _caminhos() -> list[Path]:
    """Os arquivos das calculadoras sob CDI e CDI + Spread (uma pasta por ativo)."""
    saida: list[Path] = []
    for label in ("CDI", "CDI + Spread"):
        base = HOMOLOG_ROOT / label
        if base.exists():
            saida += sorted(
                p for p in base.rglob("*")
                if p.suffix.lower() in (".xlsx", ".xlsm")
            )
    return saida


def _preencher(wb, caminhos: list[Path]) -> None:
    # aba Lista
    if "Lista" in wb.sheetnames:
        ws = wb["Lista"]
        wb.remove(ws)
    ws = wb.create_sheet("Lista", 0)
    ws["A1"] = "Caminho"
    ws["A1"].font = Font(bold=True)
    for i, p in enumerate(caminhos, start=2):
        ws.cell(i, 1, str(p))
    ws.column_dimensions["A"].width = 110

    # aba LEIA-ME
    if "LEIA-ME" in wb.sheetnames:
        wb.remove(wb["LEIA-ME"])
    info = wb.create_sheet("LEIA-ME")
    for i, linha in enumerate(INSTRUCOES, start=1):
        c = info.cell(i, 1, linha)
        if i == 1:
            c.font = Font(bold=True, size=12)
    info.column_dimensions["A"].width = 95

    # remove a 'Sheet' padrão, se existir e não for usada
    for nome in ("Sheet", "Planilha1"):
        if nome in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb[nome])


def main() -> None:
    caminhos = _caminhos()
    print(f"Calculadoras encontradas: {len(caminhos)}")

    if XLSM.exists():
        try:
            wb = openpyxl.load_workbook(str(XLSM), keep_vba=True)
        except PermissionError:
            raise SystemExit(f"Feche o arquivo no Excel antes de atualizar:\n  {XLSM}")
        _preencher(wb, caminhos)
        wb.save(str(XLSM))
        print(f"Lista atualizada (macro preservada): {XLSM}")
    else:
        wb = openpyxl.Workbook()
        _preencher(wb, caminhos)
        wb.save(str(XLSX))
        print(f"Criado: {XLSX}")
        print("Próximo passo: abra-o, cole a macro RecalcCache no VBE (Alt+F11) e")
        print("salve como 'Rotina_Recalc.xlsm'.")


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    main()

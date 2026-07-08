"""Script único de produção: CALCULA os PUs de D-0 e GERA o workbook XLSX.

Calcula (coletar_pus) e apresenta (gerar_planilha) num só lugar.

Layout do workbook:
- uma ABA por família (di_puro, di_spread, prefixado, ipca_spread), com colunas
  IdTitulo, IdSerie, Data, PU, Observação;
- uma aba "Referências" com Nome, Referência (caminho da calculadora) e Família,
  com filtro (autofilter) para organizar por família.
"""
import sys
from pathlib import Path

# Permite rodar tanto como módulo (python -m mtm_skills.gerar_planilha_pu) quanto
# direto (python mtm_skills/gerar_planilha_pu.py): garante a raiz do projeto no path.
_RAIZ = Path(__file__).resolve().parents[1]
if str(_RAIZ) not in sys.path:
    sys.path.insert(0, str(_RAIZ))

from datetime import date

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from pu_mtm.app import config
from pu_mtm.dados import csvio
from pu_mtm.dados.cadastro import ler_cadastro, ler_eventos
from pu_mtm.dados.feriados import feriados_b3
from pu_mtm.dados.indice_mercado import cdi_por_data, ipca_var_por_mes
from pu_mtm.dominio.nucleo_pu import calcular_pu, arred_juros_por_nome
from pu_mtm.dominio.familias.di import fator_juros_acumulado
from pu_mtm.dominio.familias import prefixado as fam_prefixado
from pu_mtm.dominio.familias import ipca as fam_ipca
from pu_mtm.dados.grade import montar_dias_prefixado, montar_dias_ipca
from pu_mtm.app.rodar_verificacao import montar_dias

# Regra 6 (Congelados): PU fixo (sem accrual), inserido à mão, não vinculado à
# calculadora. Só incluir o ativo aqui quando o valor congelado for conhecido —
# nunca fabricar um chute. Mecanismo mantido para futuros congelados de verdade.
# (464/474 NÃO são congelados: estão ativos a 105% do CDI e o motor bate a PU_MtM
#  no dígito — ex.: 464=0.0134597053 em 2026-06-17 = valor da mestre.)
VALORES_CONGELADOS = {}

# Calculadoras com divergência conhecida (motor correto, planilha com bug não corrigido).
# 638/743/750/762 foram corrigidas em 2026-06-25 — removidas daqui (ver docs/bugs-calculadoras.md).
# 304 (Ativo excluido) excluída do cadastro por decisão; não entra no refresh diário.
BUGS_CALCULADORA: dict[str, str] = {}

_MESES_PT = ("Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
             "Jul", "Ago", "Set", "Out", "Nov", "Dez")

def _saida_dir(d: date) -> Path:
    return (config.SAIDA_ROOT
            / f"Saída ({d.year})"
            / f"Saída - {_MESES_PT[d.month - 1]}.{str(d.year)[2:]}")

FAMILIAS = ["di_puro", "di_spread", "prefixado", "ipca_spread"]
COLUNAS_FAMILIA = ["IdTitulo", "IdSerie", "Data", "PU", "Observação"]
COLUNAS_REFERENCIAS = ["Nome", "Referência", "Família"]


# --------------------------------------------------------------------------- #
#  CÁLCULO
# --------------------------------------------------------------------------- #
def coletar_pus(data_ref: date = None) -> list[dict]:
    """Calcula o PU de D-0 de cada ativo não-IPCA do cadastro, aplicando status
    (liquidado→ignora; congelado→PU fixo) e marcando calculadoras com bug.
    Retorna registros: {id_serie, id_titulo, familia, data, pu, observacao}."""
    if data_ref is None:
        data_ref = date.today()

    ativos = ler_cadastro(str(config.CADASTRO))
    # Status e IdTitulo são dados cadastrais (não da matemática do Ativo): lê do CSV cru.
    linhas_cadastro = list(csvio.ler_dict(str(config.CADASTRO)))
    status_por_ativo = {l["IdSerie"]: str(l.get("Status", "")).lower() for l in linhas_cadastro}
    idtitulo_por_ativo = {l["IdSerie"]: str(l.get("IdTitulo", "")) for l in linhas_cadastro}

    # Carrega Feriados e CDI uma única vez para todos os ativos
    feriados = feriados_b3(2000, data_ref.year + 1)
    tem_ativos_di = any(a.familia in ("di_puro", "di_spread") for a in ativos.values())
    cdi_map = cdi_por_data(str(config.BASE_MERCADO), serie=config.SERIE_DI) if tem_ativos_di else {}
    tem_ativos_ipca = any(a.familia == "ipca_spread" for a in ativos.values())
    ipca_var = ipca_var_por_mes(str(config.BASE_MERCADO)) if tem_ativos_ipca else {}

    # Não dá para marcar ativo DI num dia sem CDI: ancora o D-0 ao último dia de
    # mercado disponível (a curva pode não ter o dia de hoje se a base ainda não
    # foi atualizada). Mantém todas as famílias na mesma data de referência.
    if cdi_map:
        ultima_mercado = max(cdi_map)
        if data_ref > ultima_mercado:
            print(f"Aviso: CDI disponível até {ultima_mercado}; ancorando D-0 nessa data (hoje={data_ref}).")
            data_ref = ultima_mercado

    def _rec(ativo, pu, obs=""):
        return {"id_serie": ativo.id_serie, "id_titulo": idtitulo_por_ativo.get(ativo.id_serie, ""),
                "familia": ativo.familia, "data": data_ref, "pu": pu, "observacao": obs}

    registros = []
    for id_serie, ativo in ativos.items():
        status = status_por_ativo.get(id_serie, "")

        # --- Tratamento de Status (Regra 6 do Handoff) ---
        if "liquidad" in status:
            continue                                   # liquidado (ex.: 512): ignora

        if "congelad" in status:
            # Congelado: PU fixo (sem accrual). Só exporta se o valor for conhecido —
            # não fabrica 0.0 nem chute para os que faltam.
            if id_serie in VALORES_CONGELADOS:
                registros.append(_rec(ativo, VALORES_CONGELADOS[id_serie], "congelado (PU fixo)"))
            else:
                print(f"Congelado sem PU fixo cadastrado, pulado: {id_serie} ({ativo.apelido})")
            continue

        # --- Pipeline Normal de Cálculo ---
        eventos = ler_eventos(str(config.EVENTOS_DIR), id_serie, ativo.familia)
        try:
            if ativo.familia == "prefixado":
                dias = montar_dias_prefixado(ativo, data_ref, eventos, feriados)
                fator_fn = fam_prefixado.fator_juros_acumulado
            elif ativo.familia == "ipca_spread":
                dias = montar_dias_ipca(ativo, data_ref, eventos, ipca_var, feriados)
                fator_fn = fam_ipca.fator_juros_acumulado
            else:
                dias = montar_dias(ativo, data_ref, feriados, cdi_map)
                fator_fn = fator_juros_acumulado
            r = calcular_pu(ativo, dias, eventos, fator_fn, arred_juros_por_nome(ativo.juros_arred))
            registros.append(_rec(ativo, r.pu, BUGS_CALCULADORA.get(id_serie, "")))
        except Exception as e:
            print(f"Erro ao calcular PU do ativo {id_serie} ({ativo.apelido}): {e}")

    return registros


# --------------------------------------------------------------------------- #
#  GERAÇÃO DO WORKBOOK
# --------------------------------------------------------------------------- #
def _formatar(ws, ncols: int):
    """Cabeçalho em negrito, painel congelado, autofiltro e largura automática."""
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max(ws.max_row, 1)}"
    for col in range(1, ncols + 1):
        letra = get_column_letter(col)
        larg = max((len(str(c.value)) for c in ws[letra] if c.value is not None), default=10)
        ws.column_dimensions[letra].width = min(max(larg + 2, 12), 70)


def gerar_planilha(data_ref: date = None, caminho_saida: str = None) -> str:
    """Calcula (coletar_pus) e escreve o workbook XLSX. Por padrão grava em
    `data/Saída/Saída (YYYY)/Saída - MMM.YY/saida_pu_YYYY-MM-DD.xlsx`
    onde a data é a data efetiva dos PUs (D-1 quando o CDI de hoje não existe)."""
    registros = coletar_pus(data_ref)
    if caminho_saida is None:
        data_ef = registros[0]["data"] if registros else date.today()
        saida_dir = _saida_dir(data_ef)
        saida_dir.mkdir(parents=True, exist_ok=True)
        caminho_saida = str(saida_dir / f"saida_pu_{data_ef.isoformat()}.xlsx")

    por_familia = {fam: [] for fam in FAMILIAS}
    for r in registros:
        por_familia.setdefault(r["familia"], []).append(r)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove a aba default

    # 1) uma aba por família
    for fam in FAMILIAS:
        ws = wb.create_sheet(title=fam)
        ws.append(COLUNAS_FAMILIA)
        for r in sorted(por_familia.get(fam, []), key=lambda x: x["id_serie"]):
            ws.append([r["id_titulo"], r["id_serie"], r["data"].isoformat(),
                       r["pu"], r["observacao"]])
        _formatar(ws, len(COLUNAS_FAMILIA))

    # 2) aba de Referências (Nome, Referência, Família) — com filtro
    ws = wb.create_sheet(title="Referências")
    ws.append(COLUNAS_REFERENCIAS)
    for l in csvio.ler_dict(str(config.CADASTRO)):
        ws.append([l.get("Apelido", ""), l.get("CalcPath", ""), l.get("Familia", "")])
    _formatar(ws, len(COLUNAS_REFERENCIAS))

    Path(caminho_saida).parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(caminho_saida)
    except PermissionError:
        raise SystemExit(
            f"Não consegui gravar '{caminho_saida}': o arquivo está aberto no Excel.\n"
            "Feche a planilha de saída (saida_pu.xlsx) e rode de novo.")
    n_fam = sum(1 for f in FAMILIAS if por_familia.get(f))
    print(f"Planilha gerada em: {caminho_saida} | {len(registros)} PUs em {n_fam} família(s).")
    return caminho_saida


if __name__ == "__main__":
    gerar_planilha()

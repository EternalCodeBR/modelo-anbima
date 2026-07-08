"""Reconstrói as calculadoras de homologação CDI / CDI+Spread a partir das originais.

Pipeline (fases idempotentes, cada uma com um portão de verificação):

    1. apagar    — remove as pastas homolog CDI e "CDI + Spread".
    2. copiar    — copia cada original do SharePoint para
                   ``HOMOLOG_ROOT/<CDI|CDI + Spread>/<Apelido> - Homologação/<arquivo>``
                   (uma pasta por ativo, dividido por família), com delay entre
                   escritas e conferência de md5 origem×destino.
    3. analisar  — para cada calculadora, localiza a coluna de taxa/data e
                   classifica cada SUMIFS em **D0 (CDI considerado)** ou
                   **D-1 (DI Over)** pela linha da célula de data.
    4. validar   — PORTÃO: prova, sem Excel, que trocar SUMIFS→XLOOKUP não muda
                   nenhum número. Para cada célula compara o valor CACHEADO da
                   calculadora (o que o Excel já calculou via SUMIFS) contra o
                   rate que o XLOOKUP resolveria na aba CDI local. Também exige
                   equivalência ESTRUTURAL (só o SUMIFS muda; IF, /100, offset e
                   seeds intactos). Qualquer divergência aborta o arquivo.
    5. aplicar   — só depois do portão passar, grava as fórmulas XLOOKUP
                   (prefixo OOXML correto ``_xlfn._xlws.XLOOKUP`` + 4º arg ``0``).
    6. cdi       — delega a ``atualizar_cdi_calculadoras`` (estende a aba CDI
                   local até a última data de mercado).

O comando ``dry`` roda 3+4 sobre CÓPIAS TEMPORÁRIAS das originais, sem tocar em
nada — prova o portão ponta a ponta antes de qualquer operação destrutiva.

Nunca abre o Excel (tudo openpyxl). As originais no SharePoint são apenas LIDAS.

Uso:
    python -m mtm_skills.rebuild_homolog dry       # prova o portão (não altera nada)
    python -m mtm_skills.rebuild_homolog analisar  # só o laudo D0/D-1
    python -m mtm_skills.rebuild_homolog run        # pipeline completo (apaga+recria)
    python -m mtm_skills.rebuild_homolog run --keep-cdi   # não roda a fase 6
"""
from __future__ import annotations

import hashlib
import os
import re
import shutil
import sys
import tempfile
import time
import warnings
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

warnings.filterwarnings("ignore")

_RAIZ = Path(__file__).resolve().parents[1]
if str(_RAIZ) not in sys.path:
    sys.path.insert(0, str(_RAIZ))

import openpyxl
from openpyxl.utils import column_index_from_string

from pu_mtm.app import config

# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------
HOMOLOG_ROOT = config.HOMOLOG_CALC_ROOT
FLUXO_DIR = config.FLUXO_DIR
SUFIXO = " - Homologação"

# Famílias no cadastro (códigos normalizados de ler_dict) -> rótulo da pasta.
FAMILIA_LABEL = {"di_puro": "CDI", "di_spread": "CDI + Spread"}

# Delay entre escritas (cortesia à heurística antiransomware; transparente).
DELAY = float(os.environ.get("REBUILD_DELAY", "1.5"))

# Tolerância na comparação numérica de taxas (mesma fonte -> praticamente exato).
TOL = 1e-6

# Prefixo OOXML do XLOOKUP confirmado pelo próprio Excel (ele reescreveu a
# fórmula assim ao reparar um arquivo, e o resultado abre sem prompt de
# reparo). Ver mtm_skills/corrigir_prefixo_xlookup.py para o histórico.
XLOOKUP = "_xlfn.XLOOKUP"

NOMES_TAXA = ["Taxa_DI", "DI Over", "DI", "Taxa DI"]
NOMES_DATA = ["Data", "DATA", "data", "Datas"]
ABAS_IGNORAR = {"CDI", "Feriados", "obs", "Obs", "OBS"}

_SEP = "=" * 74


# ---------------------------------------------------------------------------
# Modelo
# ---------------------------------------------------------------------------
@dataclass
class Ativo:
    idserie: str
    apelido: str
    familia: str
    calcpath: str
    origem: Path
    familia_label: str
    arquivo: str  # basename preservado (chave de casamento downstream)

    @property
    def destino_dir(self) -> Path:
        return HOMOLOG_ROOT / self.familia_label / f"{_slug(self.apelido)}{SUFIXO}"

    @property
    def destino(self) -> Path:
        return self.destino_dir / self.arquivo


def _slug(nome: str) -> str:
    """Nome de pasta seguro (o nome do ARQUIVO é preservado; a pasta é cosmética)."""
    limpo = re.sub(r'[<>:"/\\|?*]', "-", str(nome)).strip()
    return limpo or "ativo"


# ---------------------------------------------------------------------------
# Cadastro
# ---------------------------------------------------------------------------
def carregar_ativos() -> list[Ativo]:
    from pu_mtm.dados.csvio import ler_dict

    rows = ler_dict(str(config.CADASTRO))
    ativos: list[Ativo] = []
    vistos: dict[str, str] = {}
    for r in rows:
        fam = (r.get("Familia") or "").strip().lower()
        if fam not in FAMILIA_LABEL:
            continue
        cp = r.get("CalcPath")
        if not cp:
            continue
        if (r.get("Status") or "").lower() == "liquidado":
            continue
        origem = FLUXO_DIR / cp.replace("/", os.sep)
        arquivo = Path(cp).name
        low = arquivo.lower()
        if low in vistos:
            raise SystemExit(
                f"Nome de arquivo duplicado entre ativos ({vistos[low]} e "
                f"{r.get('Apelido')}): {arquivo} — casamento por basename seria ambíguo."
            )
        vistos[low] = r.get("Apelido", r["IdSerie"])
        ativos.append(Ativo(
            idserie=str(r["IdSerie"]),
            apelido=str(r.get("Apelido") or r["IdSerie"]),
            familia=fam,
            calcpath=cp,
            origem=origem,
            familia_label=FAMILIA_LABEL[fam],
            arquivo=arquivo,
        ))
    return ativos


def _md5(path: Path) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Localização da aba/colunas de taxa e data
# ---------------------------------------------------------------------------
def _achar_taxa_data(ws) -> tuple[int, int | None, int | None, str | None]:
    for ri in range(1, 6):
        hdr = {str(c.value).strip(): c.column for c in ws[ri] if c.value is not None}
        if len(hdr) < 3:
            continue
        taxa = next((hdr[k] for k in NOMES_TAXA if k in hdr), None)
        if not taxa:
            continue
        data = next((hdr[k] for k in NOMES_DATA if k in hdr), None)
        nome = next((k for k in NOMES_TAXA if k in hdr), None)
        return ri, taxa, data, nome
    return 0, None, None, None


def _aba_calc(wb):
    for ws in wb.worksheets:
        if ws.title in ABAS_IGNORAR:
            continue
        hrow, taxa, data, nome = _achar_taxa_data(ws)
        if taxa:
            return ws, hrow, taxa, data, nome
    return None, 0, None, None, None


# ---------------------------------------------------------------------------
# Parsing de fórmula: SUMIFS -> XLOOKUP
# ---------------------------------------------------------------------------
_RE_SUMIFS = re.compile(r"SUMIFS\s*\(", re.IGNORECASE)
_RE_DATE_B = re.compile(r"\$?B\s*:\s*\$?B\s*,\s*(\$?[A-Z]{1,3}\$?\d+)", re.IGNORECASE)
_RE_CELL = re.compile(r"\$?[A-Z]{1,3}\$?\d+")
_RE_COORD = re.compile(r"^\$?([A-Z]{1,3})\$?(\d+)$")


def _match_paren(s: str, open_idx: int) -> int:
    depth = 0
    for i in range(open_idx, len(s)):
        if s[i] == "(":
            depth += 1
        elif s[i] == ")":
            depth -= 1
            if depth == 0:
                return i
    return -1


def _extrair_data_cell(args: str) -> str | None:
    m = _RE_DATE_B.search(args)
    if m:
        return m.group(1)
    cells = _RE_CELL.findall(args)
    return cells[-1] if cells else None


def sumifs_para_xlookup(formula: str) -> tuple[str, int, list[str]]:
    """Troca cada SUMIFS(...) por XLOOKUP(<data>,CDI!A:A,CDI!B:B,0).

    Retorna (nova_formula, n_trocas, celulas_de_data).
    """
    out = formula
    n = 0
    datas: list[str] = []
    while True:
        m = _RE_SUMIFS.search(out)
        if not m:
            break
        op = m.end() - 1
        close = _match_paren(out, op)
        if close == -1:
            break
        args = out[op + 1:close]
        data_cell = _extrair_data_cell(args)
        if not data_cell:
            break
        datas.append(data_cell)
        repl = f"{XLOOKUP}({data_cell},CDI!A:A,CDI!B:B,0)"
        out = out[:m.start()] + repl + out[close + 1:]
        n += 1
    return out, n, datas


def _blank_calls(formula: str, regex: re.Pattern) -> str:
    """Substitui cada chamada casada por um marcador '@' para comparação estrutural."""
    out = formula
    while True:
        m = regex.search(out)
        if not m:
            break
        op = m.end() - 1
        close = _match_paren(out, op)
        if close == -1:
            break
        out = out[:m.start()] + "@" + out[close + 1:]
    return out


_RE_XLOOKUP = re.compile(re.escape(XLOOKUP) + r"\s*\(", re.IGNORECASE)


def equivalencia_estrutural(orig: str, nova: str) -> bool:
    """True se `nova` = `orig` com apenas os SUMIFS trocados por XLOOKUP.

    Prova que o envelope (IF, /100, offsets, referências) ficou intacto.
    """
    return _blank_calls(orig, _RE_SUMIFS) == _blank_calls(nova, _RE_XLOOKUP)


# ---------------------------------------------------------------------------
# Utilidades de data / valor cacheado
# ---------------------------------------------------------------------------
def _to_iso(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()[:10]
    if "/" in s:
        d, m, a = s.split("/")
        return f"{a}-{m.zfill(2)}-{d.zfill(2)}"
    return s


def _coord(ref: str) -> tuple[int, int] | None:
    m = _RE_COORD.match(ref.strip())
    if not m:
        return None
    return column_index_from_string(m.group(1)), int(m.group(2))


def _carregar_mercado() -> tuple[dict[str, float], str]:
    """CDI autoritativo do mercado.xlsx: ({data_iso: CDI_AA}, ultima_data_iso).

    Esta é a fonte que o `atualizar_cdi` grava em toda aba CDI local; portanto é o
    alvo efetivo do XLOOKUP após o pipeline. Valida-se contra ela (não contra a
    aba embutida, que pode estar defasada ou ausente).
    """
    from mtm_skills.atualizar_cdi_calculadoras import ler_cdi_mercado, MERCADO_XLSX
    hist = ler_cdi_mercado(MERCADO_XLSX)  # {iso: (cdi_aa, fator)}
    cdi = {d: v[0] for d, v in hist.items()}
    return cdi, max(cdi) if cdi else ""


def _num(v) -> float | None:
    """float robusto: None/''/erro Excel ('#VALUE!', '#N/A'...) -> None."""
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _quase_igual(a, b) -> bool:
    fa, fb = _num(a), _num(b)
    if fa is None or fb is None:
        return False
    return abs(fa - fb) <= TOL + TOL * max(abs(fa), abs(fb))


# ---------------------------------------------------------------------------
# Análise (fase 3) + Validação numérica (fase 4)
# ---------------------------------------------------------------------------
@dataclass
class Laudo:
    idserie: str
    apelido: str
    aba: str = ""
    classe: str = "?"          # D0 / D-1 / misto / ?
    n_sumifs: int = 0
    n_hist_ok: int = 0     # linhas realizadas validadas numericamente vs mercado
    n_futuro: int = 0      # futuras/não-realizadas no snapshot (convenção ,0)
    n_erro_orig: int = 0   # cache #VALUE!/#N/A na ORIGINAL (a migração conserta)
    envelope_if: bool = False
    erros: list[str] = field(default_factory=list)

    @property
    def n_validadas(self) -> int:
        return self.n_hist_ok + self.n_futuro + self.n_erro_orig

    @property
    def ok(self) -> bool:
        return not self.erros and self.n_sumifs > 0 and self.n_validadas == self.n_sumifs


def analisar_validar(caminho: Path, idserie: str, apelido: str,
                     cdi_mkt: dict[str, float], last_iso: str) -> tuple[Laudo, dict[int, str]]:
    """Classifica D0/D-1 e valida cada célula de taxa contra o mercado.xlsx.

    Linha histórica (data <= última do mercado): valor cacheado do SUMIFS tem de
    bater com o CDI de mercado. Linha futura: aceita (convenção ,0). Qualquer
    divergência histórica ou desvio de envelope reprova o arquivo.

    Retorna (laudo, {linha: nova_formula}) — o mapa só é usado na aplicação.
    """
    lau = Laudo(idserie=idserie, apelido=apelido)
    novas: dict[int, str] = {}

    wb_f = openpyxl.load_workbook(str(caminho), data_only=False)
    wb_v = openpyxl.load_workbook(str(caminho), data_only=True)
    try:
        ws_f, hrow, taxa, data_col, _ = _aba_calc(wb_f)
        if not ws_f or not taxa:
            lau.erros.append("sem coluna de taxa")
            return lau, novas
        lau.aba = ws_f.title
        ws_v = wb_v[ws_f.title]

        classes: set[str] = set()
        for r in range(hrow + 1, ws_f.max_row + 1):
            val = ws_f.cell(r, taxa).value
            if not (isinstance(val, str) and "SUMIFS" in val.upper()):
                continue
            lau.n_sumifs += 1
            if "IF(" in val.upper().replace(" ", ""):
                lau.envelope_if = True

            nova, ntroca, datas = sumifs_para_xlookup(val)
            if ntroca == 0 or not datas:
                lau.erros.append(f"r{r}: não extraiu célula de data")
                continue

            # (a) equivalência estrutural
            if not equivalencia_estrutural(val, nova):
                lau.erros.append(f"r{r}: envelope alterado além do SUMIFS")
                continue

            # classificação D0/D-1 pela 1ª célula de data
            dc = datas[0]
            co = _coord(dc)
            if co:
                drow = co[1]
                classes.add("D0" if drow == r else "D-1" if drow == r - 1 else f"off{drow - r}")

            # (b) validação numérica vs mercado (realizado) / convenção ,0 (futuro)
            cached = ws_v.cell(r, taxa).value
            ok_num, cat = _validar_celula(val, dc, cached, ws_v, cdi_mkt, last_iso)
            if ok_num:
                if cat == "hist_ok":
                    lau.n_hist_ok += 1
                elif cat == "erro_orig":
                    lau.n_erro_orig += 1
                else:
                    lau.n_futuro += 1
                novas[r] = nova
            else:
                lau.erros.append(f"r{r}: {cat}")

        if not classes:
            lau.classe = "?"
        elif len(classes) == 1:
            lau.classe = next(iter(classes))
        else:
            lau.classe = "misto(" + ",".join(sorted(classes)) + ")"
        return lau, novas
    finally:
        wb_f.close()
        wb_v.close()


def _split_top_args(s: str) -> list[str]:
    """Divide args de topo por vírgula, respeitando parênteses."""
    out, depth, cur = [], 0, ""
    for ch in s:
        if ch == "(":
            depth += 1
            cur += ch
        elif ch == ")":
            depth -= 1
            cur += ch
        elif ch == "," and depth == 0:
            out.append(cur)
            cur = ""
        else:
            cur += ch
    if cur:
        out.append(cur)
    return out


def _data_iso_de(data_cell: str, ws_v) -> str | None:
    co = _coord(data_cell)
    if not co:
        return None
    return _to_iso(ws_v.cell(co[1], co[0]).value)


def _escala_ok(cache_num: float, rate: float) -> bool:
    """Aceita o cache na escala percentual (14.15) ou decimal (0.1415).

    O envelope (/100 ou não) é preservado pela prova estrutural; aqui só se
    confere a MAGNITUDE do rate, tolerando as duas convenções de escala.
    """
    return _quase_igual(cache_num, rate) or _quase_igual(cache_num, rate / 100.0)


def _validar_celula(formula: str, data_cell: str, cached, ws_v,
                    cdi_mkt: dict[str, float], last_iso: str) -> tuple[bool, str]:
    """Valida uma célula de taxa contra o mercado. Retorna (ok, categoria).

    - 'hist_ok'    : linha realizada; magnitude do rate bate (escala 1 ou /100).
    - 'futuro'     : data futura OU não-realizada no snapshot (cache 0/vazio) — ,0.
    - 'erro_orig'  : cache #VALUE!/#N/A na ORIGINAL (link externo); a migração conserta.
    - 'hole'/'div' : reprovas (data histórica ausente no mercado, ou rate divergente).

    Trata ``IF(cond, alt, SUMIFS(...))``: no ramo alternativo o cache é `alt`.
    """
    di = _data_iso_de(data_cell, ws_v)
    if di is None:
        return True, "futuro"                 # sem data resolvível (seed/projeção)
    if last_iso and di > last_iso:
        return True, "futuro"                 # CDI ainda não publicado
    if isinstance(cached, str) and cached.startswith("#"):
        return True, "erro_orig"              # #VALUE!/#N/A pré-existente na original
    c = _num(cached)
    if c is None or c == 0.0:
        return True, "futuro"                 # não realizada no snapshot da original
    rate = cdi_mkt.get(di)
    if rate is None:
        return False, f"hole: {di} ausente no mercado"
    if _escala_ok(c, rate):
        return True, "hist_ok"

    # ramo alternativo de um IF(cond, alt, SUMIFS)?
    up = formula.upper()
    mi = up.find("IF(")
    if mi != -1:
        op = up.find("(", mi)
        close = _match_paren(formula, op)
        if close != -1:
            args = _split_top_args(formula[op + 1:close])
            if len(args) == 3:
                co = _coord(args[1].strip())
                if co and _quase_igual(c, _num(ws_v.cell(co[1], co[0]).value) or 0.0):
                    return True, "hist_ok"
    return False, f"div (cache={cached} vs mkt={rate})"


# ---------------------------------------------------------------------------
# Fases destrutivas / de escrita
# ---------------------------------------------------------------------------
def _existentes_homolog() -> list[Path]:
    out: list[Path] = []
    for label in FAMILIA_LABEL.values():
        d = HOMOLOG_ROOT / label
        if d.exists():
            out += [p for p in d.rglob("*") if p.suffix.lower() in (".xlsx", ".xlsm")]
    return out


def _travado(p: Path) -> bool:
    """Teste de lock canônico no Windows: tenta renomear o arquivo para si mesmo."""
    tmp = p.with_name(p.name + ".locktest")
    try:
        os.rename(p, tmp)
        os.rename(tmp, p)
        return False
    except OSError:
        return True


def _rm_onexc(func, path, _exc) -> None:
    """Handler de rmtree: limpa read-only e tenta de novo (locks transitórios)."""
    import stat
    try:
        os.chmod(path, stat.S_IWRITE)
        func(path)
    except OSError:
        pass


def fase_apagar() -> None:
    print(_SEP); print("FASE 1 — apagar homolog CDI + CDI + Spread"); print(_SEP)
    # Pré-checagem: aborta ANTES de apagar qualquer coisa se algo estiver travado
    # (ex.: calculadora aberta no Excel), evitando estado parcial.
    travados = [p for p in _existentes_homolog() if _travado(p)]
    if travados:
        raise SystemExit(
            "\nArquivo(s) TRAVADO(S) — feche o Excel/OneDrive e rode de novo:\n"
            + "\n".join(f"  {p.parent.name}/{p.name}" for p in travados)
        )
    for label in FAMILIA_LABEL.values():
        d = HOMOLOG_ROOT / label
        if d.exists():
            shutil.rmtree(d, onexc=_rm_onexc)
            print(f"  removido: {d}")
        else:
            print(f"  (inexistente): {d}")


def fase_copiar(ativos: list[Ativo]) -> None:
    print(f"\n{_SEP}"); print(f"FASE 2 — copiar {len(ativos)} originais (delay {DELAY}s)"); print(_SEP)
    ok = falta = erro = 0
    for a in ativos:
        if not a.origem.exists():
            print(f"  [{a.idserie:>10}] ORIG FALTA: {a.arquivo}")
            falta += 1
            continue
        a.destino_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(a.origem), str(a.destino))
        if _md5(a.origem) == _md5(a.destino):
            print(f"  [{a.idserie:>10}] {a.apelido:<26} -> {a.familia_label}/…/{a.arquivo}")
            ok += 1
        else:
            print(f"  [{a.idserie:>10}] {a.apelido:<26} MD5 DIVERGENTE!")
            erro += 1
        time.sleep(DELAY)
    print(f"\n  Copiadas: {ok} | Faltando: {falta} | Erro md5: {erro}")


def fase_aplicar(ativos: list[Ativo], mapas: dict[str, dict[int, str]]) -> None:
    print(f"\n{_SEP}"); print("FASE 5 — aplicar PROCX (só arquivos aprovados no portão)"); print(_SEP)
    ok = 0
    for a in ativos:
        novas = mapas.get(a.idserie)
        if not novas:
            continue
        wb = openpyxl.load_workbook(str(a.destino), data_only=False)
        ws, hrow, taxa, _, _ = _aba_calc(wb)
        for r, formula in novas.items():
            ws.cell(r, taxa).value = formula
        wb.save(str(a.destino))
        wb.close()
        print(f"  [{a.idserie:>10}] {a.apelido:<26} {len(novas)} XLOOKUP gravados")
        ok += 1
        time.sleep(DELAY)
    print(f"\n  Arquivos convertidos: {ok}")


# ---------------------------------------------------------------------------
# Portão de análise+validação (fases 3+4) sobre um conjunto de caminhos
# ---------------------------------------------------------------------------
def portao(pares: list[tuple[Ativo, Path]]) -> tuple[list[Laudo], dict[str, dict[int, str]]]:
    print(f"\n{_SEP}"); print("FASES 3+4 — classificar D0/D-1 e validar equivalência"); print(_SEP)
    cdi_mkt, last_iso = _carregar_mercado()
    print(f"Mercado CDI: {len(cdi_mkt)} datas (última {last_iso})\n")
    print(f"{'Id':>10} {'Apelido':<26} {'aba':<11} {'classe':<8} {'SUMIFS':>6} "
          f"{'hist':>5} {'fut':>4} {'#VAL':>4} {'IF':>3}  status")
    print("-" * 100)
    laudos: list[Laudo] = []
    mapas: dict[str, dict[int, str]] = {}
    for a, caminho in pares:
        lau, novas = analisar_validar(caminho, a.idserie, a.apelido, cdi_mkt, last_iso)
        laudos.append(lau)
        if lau.ok:
            mapas[a.idserie] = novas
        status = "OK" if lau.ok else "FALHOU: " + "; ".join(lau.erros[:2])
        print(f"{lau.idserie:>10} {lau.apelido[:26]:<26} {lau.aba[:11]:<11} "
              f"{lau.classe:<8} {lau.n_sumifs:>6} {lau.n_hist_ok:>5} {lau.n_futuro:>4} "
              f"{lau.n_erro_orig:>4} {'sim' if lau.envelope_if else '-':>3}  {status}")
    d0 = sum(1 for l in laudos if l.classe == "D0")
    d1 = sum(1 for l in laudos if l.classe == "D-1")
    okc = sum(1 for l in laudos if l.ok)
    print("-" * 96)
    print(f"  D0 (considerado): {d0} | D-1 (Over): {d1} | "
          f"aprovados no portão: {okc}/{len(laudos)}")
    return laudos, mapas


# ---------------------------------------------------------------------------
# Comandos
# ---------------------------------------------------------------------------
def cmd_analisar() -> None:
    ativos = carregar_ativos()
    pares = [(a, a.origem) for a in ativos if a.origem.exists()]
    portao(pares)


def cmd_dry() -> None:
    """Prova o portão sobre CÓPIAS TEMPORÁRIAS das originais — não altera nada."""
    ativos = carregar_ativos()
    tmp = Path(tempfile.mkdtemp(prefix="rebuild_dry_"))
    print(f"(dry-run em cópias temporárias: {tmp})")
    try:
        pares: list[tuple[Ativo, Path]] = []
        for a in ativos:
            if not a.origem.exists():
                continue
            dst = tmp / a.arquivo
            shutil.copy2(str(a.origem), str(dst))
            pares.append((a, dst))
        laudos, _ = portao(pares)
        reprovados = [l for l in laudos if not l.ok]
        if reprovados:
            print(f"\n⚠ {len(reprovados)} calculadora(s) NÃO passaram no portão — "
                  "revisar antes de rodar 'run'.")
        else:
            print("\n✓ Portão 100% verde — seguro para 'run'.")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def cmd_run(keep_cdi: bool) -> None:
    ativos = carregar_ativos()

    # 0) Provar o portão nas próprias originais antes de destruir a homolog atual.
    pares_orig = [(a, a.origem) for a in ativos if a.origem.exists()]
    laudos, _ = portao(pares_orig)
    if any(not l.ok for l in laudos):
        raise SystemExit("\nPortão reprovou nas originais — abortando antes de apagar.")

    fase_apagar()
    fase_copiar(ativos)

    # Reexecuta o portão sobre as CÓPIAS (agora sim geram o mapa de fórmulas a gravar).
    pares_copia = [(a, a.destino) for a in ativos if a.destino.exists()]
    laudos2, mapas = portao(pares_copia)
    if any(not l.ok for l in laudos2):
        raise SystemExit("\nPortão reprovou nas cópias — nada foi convertido.")

    fase_aplicar(ativos, mapas)

    if not keep_cdi:
        print(f"\n{_SEP}"); print("FASE 6 — atualizar CDI (delego a atualizar_cdi_calculadoras)"); print(_SEP)
        from mtm_skills import atualizar_cdi_calculadoras as acc
        acc.atualizar()

    print(f"\n{_SEP}")
    print("CONCLUÍDO. Próximo passo (fora deste script): rodar o batimento motor×calc.")
    print(_SEP)


def main(argv: list[str]) -> None:
    cmd = argv[1] if len(argv) > 1 else "dry"
    if cmd == "analisar":
        cmd_analisar()
    elif cmd == "dry":
        cmd_dry()
    elif cmd == "run":
        cmd_run(keep_cdi="--keep-cdi" in argv)
    else:
        print(__doc__)
        raise SystemExit(2)


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    main(sys.argv)

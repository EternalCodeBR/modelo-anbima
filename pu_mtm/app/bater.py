"""Batimento genérico motor x calculadora para qualquer ativo di_puro do cadastro.

Lê do cadastro: CalcPath (relativo ao FLUXO_DIR ou absoluto), AbaPU, CelulaPU e
DataCol (coluna da data; default C). O motor calcula o PU pelo cadastro + curva DI
(via rodar_verificacao.calcular_pu_piloto); aqui só lemos o PU exibido pela planilha
e comparamos no centavo, em datas reais ao longo da vida do ativo.

Resultados de cada run são gravados em data/batimento/{id_serie}.csv (sobrescreve).

Uso:
  python -m pu_mtm.app.bater 531            # um ativo
  python -m pu_mtm.app.bater 531 530 347    # vários
"""
import sys
import warnings
from datetime import date
from pathlib import Path

warnings.filterwarnings("ignore")
import holidays
import openpyxl
from openpyxl.utils import column_index_from_string as colidx

# Calendário B3/BM&FBOVESPA (padrão do projeto — nunca derivar feriado de planilha).
_BVMF = holidays.financial_holidays("BVMF")


def _eh_dia_util(d: date) -> bool:
    """Dia útil de mercado: seg-sex e não-feriado BVMF. Em dia não-útil a
    calculadora acrua pelo calendário e o motor segura no último dia útil —
    comparar nesses dias gera falso 'erro' que reconverge no próximo útil."""
    return d.weekday() < 5 and d not in _BVMF

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

_RAIZ = Path(__file__).resolve().parents[2]
if str(_RAIZ) not in sys.path:
    sys.path.insert(0, str(_RAIZ))

from pu_mtm.app import config
from pu_mtm.app.rodar_verificacao import calcular_pu_piloto
from pu_mtm.dados import csvio
from pu_mtm.dados.cadastro import ler_cadastro
from pu_mtm.dados.indice_mercado import cdi_por_data
from pu_mtm.verificacao.comparador import comparar

_MESES_PT = ("Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
             "Jul", "Ago", "Set", "Out", "Nov", "Dez")

def _bat_dir(d: date) -> Path:
    return (config.BATIMENTO_ROOT
            / f"Batimento ({d.year})"
            / f"Batimento - {_MESES_PT[d.month - 1]}.{str(d.year)[2:]}")

_HEADER = ["DATA", "NOME", "IDTÍTULO", "CALCULADORA", "MOTOR", "DIFERENÇA"]


def _salvar_batimento(id_serie: str, apelido: str, registros: list[dict]) -> Path:
    """Grava (ou apende) em data/Batimento/Batimento (YYYY)/Batimento - MMM.YY/batimento_YYYY-MM-DD.xlsx."""
    hoje = date.today()
    out_dir = _bat_dir(hoje)
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"batimento_{hoje.isoformat()}.xlsx"
    if out.exists():
        wb = openpyxl.load_workbook(out)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(_HEADER)
    for r in registros:
        py_arred = round(r["py"], 8)
        excel_arred = round(r["excel"], 8)
        ws.append([r["data"], apelido, id_serie, excel_arred, py_arred,
                   round(py_arred - excel_arred, 8)])
    wb.save(out)
    return out


def _brl(v: float) -> str:
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _meta(id_serie: str) -> dict:
    for r in csvio.ler_dict(str(config.CADASTRO)):
        if r["IdSerie"] == id_serie:
            return r
    raise KeyError(id_serie)


def _calc_path(meta: dict) -> Path:
    # Na fase de homologação resolve para a cópia local; senão, SharePoint.
    return config.resolver_calc_path(meta["CalcPath"])


def _pu_calc(meta: dict) -> dict[date, float]:
    cam = _calc_path(meta)
    aba = meta.get("AbaPU") or "Sheet1"
    col_dt = colidx(meta.get("DataCol") or "C")
    col_pu = colidx(meta.get("CelulaPU") or "M")
    if not cam.exists():
        raise RuntimeError(f"Calculadora não encontrada: {cam}\n"
                         "Confira a coluna CalcPath do cadastro (relativa a 'Fluxo de pagamento').")
    try:
        wb = openpyxl.load_workbook(cam, data_only=True, read_only=True)
    except PermissionError:
        raise RuntimeError(f"Calculadora aberta no Excel: {cam.name}\n"
                         "Feche o arquivo (e a BaseDadosMercado) antes de auditar.")
    ws = wb[aba]
    out = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) < max(col_dt, col_pu):
            continue
        d, pu = row[col_dt - 1], row[col_pu - 1]
        d = d.date() if hasattr(d, "date") else d
        if isinstance(d, date) and isinstance(pu, (int, float)):
            out[d] = float(pu)
    wb.close()
    if not out:
        # data_only só lê o cache do Excel; sem cache (vínculo externo nunca recalculado,
        # #N/A, ou arquivo nunca salvo pelo Excel) a série vem vazia — não é erro do motor.
        raise RuntimeError(
            f"Calculadora '{cam.name}' (aba {aba}, col {meta.get('CelulaPU')}) sem PU em cache.\n"
            "Os scripts leem o ÚLTIMO valor salvo pelo Excel — não recalculam fórmula/PROCV.\n"
            "Abra a calculadora no Excel, deixe atualizar os vínculos (Dados > Atualizar Tudo),\n"
            "salve, FECHE, e rode de novo. Confira também AbaPU/CelulaPU/DataCol no cadastro.")
    return out


def _datas(pu_excel: dict[date, float], teto: date, n: int = 6) -> list[date]:
    ds = sorted(d for d in pu_excel if d <= teto and _eh_dia_util(d))
    if len(ds) <= n:
        return ds[1:]
    passos = [round(i * (len(ds) - 1) / (n - 1)) for i in range(n)]
    return sorted({ds[p] for p in passos if p > 0})


def rodar(id_serie: str, datas: list[date] | None = None) -> bool:
    meta = _meta(id_serie)
    ativo = ler_cadastro(str(config.CADASTRO))[id_serie]
    pu_excel = _pu_calc(meta)
    teto = max(cdi_por_data(str(config.BASE_MERCADO), serie=config.SERIE_DI))
    if not datas:
        datas = _datas(pu_excel, teto)

    print(f"\n  ATIVO {id_serie} - {ativo.apelido}  ({ativo.familia}, VNe R$ {_brl(ativo.vne)})")
    print("  " + "-" * 72)
    print(f"  {'Data':<12}{'PU Calculadora':>20}{'PU Motor':>20}{'Dif':>10}")
    print("  " + "-" * 72)
    todos_ok = True
    registros = []
    for d in datas:
        if d not in pu_excel or d > teto or not _eh_dia_util(d):
            continue
        pu = calcular_pu_piloto(id_serie, d)
        rel = comparar(py=pu, excel=pu_excel[d], id_serie=id_serie, vne=ativo.vne)
        marca = "OK" if rel["ok"] else "XX"
        print(f"  {d.isoformat():<12}{_brl(rel['excel']):>20}{_brl(rel['py']):>20}{marca:>10}")
        todos_ok = todos_ok and rel["ok"]
        registros.append({"data": d.isoformat(), **rel})
    print("  " + "-" * 72)
    print("  RESULTADO:", "bate no centavo." if todos_ok else "DIVERGE — ver linhas XX.")
    if registros:
        out = _salvar_batimento(id_serie, ativo.apelido, registros)
        print(f"  Planilha: {out}")
    return todos_ok


# Famílias no escopo da homologação atual: CDI (di_puro) e CDI+Spread (di_spread).
# prefixado e ipca_spread ficam FORA — senão o motor (curva DI) diverge do cache
# IPCA/prefixado da planilha e suja o batimento (ex.: 367 NC Ativo IPCA, ipca_spread).
FAMILIAS_ESCOPO = {"di_puro", "di_spread"}


def rodar_todos(familias: set[str] | None = None) -> Path | None:
    """Batimento dos ativos ativos do cadastro DENTRO do escopo de famílias.

    Por padrão restringe a CDI/CDI+Spread (``FAMILIAS_ESCOPO``). Apaga o arquivo do
    dia (se existir) antes de rodar — garante run limpo mesmo que o script seja
    executado mais de uma vez no mesmo dia.
    Retorna o caminho do arquivo gerado, ou None se nenhum ativo produziu resultado.
    """
    escopo = {f.lower() for f in (familias or FAMILIAS_ESCOPO)}
    bat_file = _bat_dir(date.today()) / f"batimento_{date.today().isoformat()}.xlsx"
    if bat_file.exists():
        bat_file.unlink()

    linhas = list(csvio.ler_dict(str(config.CADASTRO)))
    ids = [
        r["IdSerie"] for r in linhas
        if str(r.get("Familia", "")).strip().lower() in escopo
        and "liquidad" not in str(r.get("Status", "")).lower()
        and "congelad" not in str(r.get("Status", "")).lower()
        and r.get("CalcPath")
    ]

    ok_global = True
    for id_serie in ids:
        try:
            ok = rodar(id_serie)
            ok_global = ok_global and ok
        except Exception as e:
            print(f"\n  ATIVO {id_serie}: PULADO — {e}")

    return bat_file if bat_file.exists() else None


if __name__ == "__main__":
    ids = sys.argv[1:]
    if ids:
        ok = True
        for i in ids:
            try:
                ok = rodar(i) and ok
            except Exception as e:
                print(f"\n  ATIVO {i}: ERRO — {e}")
                ok = False
    else:
        ok = rodar_todos() is not None
    sys.exit(0 if ok else 1)

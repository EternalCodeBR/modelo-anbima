# pu_mtm/dados/extrair_eventos.py
"""Extrai automaticamente a agenda de eventos (Evento Juros/Amortizacao) da calculadora.
As colunas sao localizadas pelo cabecalho (rotulos universais), agnostico ao layout."""
import csv
import openpyxl
from pu_mtm.dominio.modelos import Evento

def _norm(v) -> str:
    return v.strip().lower() if isinstance(v, str) else ""

def _data(v):
    return v.date() if hasattr(v, "date") else v

def extrair_eventos(caminho_xlsx: str, aba: str, max_header: int = 12) -> list[Evento]:
    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    ws = wb[aba]
    col_data = col_ej = col_ea = None
    header_row = 1
    for r in range(1, max_header + 1):
        for c in range(1, ws.max_column + 1):
            t = _norm(ws.cell(row=r, column=c).value)
            if t == "data" and col_data is None:
                col_data, header_row = c, r
            if "evento juros" in t:
                col_ej = c
            if "evento amortiz" in t:
                col_ea = c
        if col_data and col_ej and col_ea:
            break
    if not (col_data and col_ej and col_ea):
        wb.close()
        raise ValueError(f"colunas de evento nao encontradas na aba {aba}")
    eventos = []
    for r in range(header_row + 1, ws.max_row + 1):
        d = _data(ws.cell(row=r, column=col_data).value)
        ej = ws.cell(row=r, column=col_ej).value or 0.0
        ea = ws.cell(row=r, column=col_ea).value or 0.0
        if d is not None and (ej != 0 or ea != 0):
            eventos.append(Evento(data=d, evento_juros=float(ej), evento_amortizacao=float(ea)))
    wb.close()
    return eventos

def gravar_livro(eventos: list[Evento], caminho_csv: str) -> None:
    with open(caminho_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Data", "EventoJuros", "EventoAmortizacao", "Obs"])
        for e in eventos:
            w.writerow([e.data.isoformat(), e.evento_juros, e.evento_amortizacao, "extraido do Excel"])

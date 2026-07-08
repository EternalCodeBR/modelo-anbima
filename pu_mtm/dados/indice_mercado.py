# pu_mtm/dados/indice_mercado.py
"""Leitura da curva DI e IPCA do mercado.xlsx (14-Base_Dados_Mercado).

Cada índice tem sua própria aba: CDI, FATOR_DIARIO_CDI, IPCA, etc.
Layout diário:  COD_SERIE | DATA | VALOR
Layout mensal:  COD_INDICE_MENSAL | ANO | MES | VALOR | VARIAÇÃO MENSAL
"""
from datetime import date

import openpyxl

# código de série → nome da aba no mercado.xlsx
_ABA_POR_SERIE = {
    2: "CDI",
    3: "FATOR_DIARIO_CDI",
    4: "PTAX",
    5: "EURO",
    8: "CHF",
    9: "CAD",
    11: "JPY",
}

# código de índice mensal → nome da aba
_ABA_POR_INDICE_MENSAL = {1: "IPCA"}


def _norm_data(v):
    if hasattr(v, "date"):
        return v.date()
    if isinstance(v, str):
        try:
            return date.fromisoformat(v)
        except ValueError:
            pass
    return v


def cdi_por_data(caminho_xlsx: str, serie: int, aba: str | None = None) -> dict[date, float]:
    """Retorna {date: cdi_decimal} da aba da série no mercado.xlsx.

    Valores > 1.0 estão em % a.a. (ex.: 10.4) e são divididos por 100.
    """
    nome_aba = aba or _ABA_POR_SERIE.get(serie)
    if nome_aba is None:
        raise ValueError(f"Série {serie} não mapeada para aba no mercado.xlsx")

    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True, read_only=True)
    try:
        ws = wb[nome_aba]
        out: dict[date, float] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = _norm_data(row[1])   # coluna B = DATA
            v = row[2]               # coluna C = VALOR
            if d is None or not isinstance(v, (int, float)):
                continue
            out[d] = v / 100.0 if v > 1.0 else v
        return out
    finally:
        wb.close()


def ipca_var_por_mes(caminho_xlsx: str, cod_indice: int = 1,
                     aba: str | None = None) -> dict[tuple[int, int], float]:
    """Variação mensal (%) do IPCA da aba IPCA no mercado.xlsx.

    Computa a variação em Python a partir de VALOR (índice-número) para não
    depender da coluna VARIAÇÃO MENSAL, que é fórmula Excel e pode não ter
    cache quando o arquivo foi gravado sem abrir no Excel.
    Devolve {(ano, mes): var_%}.
    """
    nome_aba = aba or _ABA_POR_INDICE_MENSAL.get(cod_indice)
    if nome_aba is None:
        raise ValueError(f"Índice mensal {cod_indice} não mapeado para aba no mercado.xlsx")

    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True, read_only=True)
    pontos: list[tuple[int, int, float]] = []
    try:
        ws = wb[nome_aba]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] != cod_indice:
                continue
            ano, mes, valor = row[1], row[2], row[3]   # ANO, MES, VALOR (índice)
            if ano is None or mes is None or valor is None:
                continue
            try:
                pontos.append((int(ano), int(mes), float(valor)))
            except (TypeError, ValueError):
                continue
    finally:
        wb.close()

    out: dict[tuple[int, int], float] = {}
    for i, (ano, mes, val) in enumerate(pontos):
        if i == 0:
            continue
        val_ant = pontos[i - 1][2]
        if val_ant:
            out[(ano, mes)] = (val / val_ant - 1) * 100
    return out
